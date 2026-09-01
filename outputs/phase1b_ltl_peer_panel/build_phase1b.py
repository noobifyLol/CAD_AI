from __future__ import annotations

import csv
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUT_DIR = Path("outputs/phase1b_ltl_peer_panel")
SRC_DIR = OUT_DIR / "sources"
XLSX = OUT_DIR / "LTL_peer_panel_PHASE1C_FINAL_2024_2026.xlsx"


SOURCES = {
    "ODFL_2024_Q1": {
        "url": "https://www.sec.gov/Archives/edgar/data/878927/000095017024047274/odfl-ex99_1.htm",
        "file": "ODFL_2024_Q1_SEC_ex99_1.htm",
        "publication_date": "2024-04-24",
        "section": "Quarterly financial and operating statistics tables",
    },
    "ODFL_2024_Q2": {
        "url": "https://www.sec.gov/Archives/edgar/data/878927/000095017024085869/odfl-ex99_1.htm",
        "file": "ODFL_2024_Q2_SEC_ex99_1.htm",
        "publication_date": "2024-07-24",
        "section": "Quarterly financial and operating statistics tables",
    },
    "ODFL_2024_Q3": {
        "url": "https://www.sec.gov/Archives/edgar/data/878927/000095017024116488/odfl-ex99_1.htm",
        "file": "ODFL_2024_Q3_SEC_ex99_1.htm",
        "publication_date": "2024-10-23",
        "section": "Quarterly financial and operating statistics tables",
    },
    "ODFL_2024_Q4_PRIOR": {
        "url": "https://www.sec.gov/Archives/edgar/data/878927/000119312526036791/odfl-ex99_1.htm",
        "file": "ODFL_2025_Q4_SEC_ex99_1.htm",
        "publication_date": "2026-02-04",
        "section": "Prior-year quarter column in 2025 Q4 comparative tables",
    },
    "ODFL_2025_Q1": {
        "url": "https://www.sec.gov/Archives/edgar/data/878927/000095017025056982/odfl-ex99_1.htm",
        "file": "ODFL_2025_Q1_SEC_ex99_1.htm",
        "publication_date": "2025-04-23",
        "section": "Quarterly financial and operating statistics tables",
    },
    "ODFL_2025_Q2": {
        "url": "https://www.sec.gov/Archives/edgar/data/878927/000095017025099899/odfl-ex99_1.htm",
        "file": "ODFL_2025_Q2_SEC_ex99_1.htm",
        "publication_date": "2025-07-23",
        "section": "Quarterly financial and operating statistics tables",
    },
    "ODFL_2025_Q3": {
        "url": "https://ir.odfl.com/sec-filings/all-sec-filings/content/0001193125-25-255040/odfl-ex99_1.htm",
        "file": "ODFL_2025_Q3_IR_SEC_ex99_1.htm",
        "publication_date": "2025-10-22",
        "section": "Quarterly financial and operating statistics tables",
    },
    "ODFL_2025_Q4": {
        "url": "https://www.sec.gov/Archives/edgar/data/878927/000119312526036791/odfl-ex99_1.htm",
        "file": "ODFL_2025_Q4_SEC_ex99_1.htm",
        "publication_date": "2026-02-04",
        "section": "Quarterly financial and operating statistics tables",
    },
    "ODFL_2026_Q1": {
        "url": "https://www.sec.gov/Archives/edgar/data/878927/000087892726000009/odfl-ex99_1.htm",
        "file": "ODFL_2026_Q1_SEC_ex99_1.htm",
        "publication_date": "2026-04-22",
        "section": "Quarterly financial and operating statistics tables",
    },
    "SAIA_2024_Q1": {
        "url": "https://www.sec.gov/Archives/edgar/data/1177702/000095017024048672/saia-ex99_1.htm",
        "file": "SAIA_2024_Q1_SEC_ex99_1.htm",
        "publication_date": "2024-04-26",
        "section": "Financial highlights and operating statistics tables",
    },
    "SAIA_2024_Q2": {
        "url": "https://www.sec.gov/Archives/edgar/data/0001177702/000095017024086727/saia-ex99_1.htm",
        "file": "SAIA_2024_Q2_SEC_ex99_1.htm",
        "publication_date": "2024-07-26",
        "section": "Financial highlights and operating statistics tables",
    },
    "SAIA_2024_Q3": {
        "url": "https://www.sec.gov/Archives/edgar/data/1177702/000095017024117217/saia-ex99_1.htm",
        "file": "SAIA_2024_Q3_SEC_ex99_1.htm",
        "publication_date": "2024-10-25",
        "section": "Financial highlights and operating statistics tables",
    },
    "SAIA_2024_Q4": {
        "url": "https://www.sec.gov/Archives/edgar/data/1177702/000095017025012023/saia-ex99_1.htm",
        "file": "SAIA_2024_Q4_SEC_ex99_1.htm",
        "publication_date": "2025-02-07",
        "section": "Financial highlights and operating statistics tables",
    },
    "SAIA_2025_Q1": {
        "url": "https://www.sec.gov/Archives/edgar/data/1177702/000095017025058331/saia-ex99_1.htm",
        "file": "SAIA_2025_Q1_SEC_ex99_1.htm",
        "publication_date": "2025-04-25",
        "section": "Financial highlights and operating statistics tables",
    },
    "SAIA_2025_Q2": {
        "url": "https://www.sec.gov/Archives/edgar/data/1177702/000095017025098577/saia-ex99_1.htm",
        "file": "SAIA_2025_Q2_SEC_ex99_1.htm",
        "publication_date": "2025-07-25",
        "section": "Financial highlights and operating statistics tables",
    },
    "SAIA_2025_Q3": {
        "url": "https://www.sec.gov/Archives/edgar/data/1177702/000119312525257441/saia-ex99_1.htm",
        "file": "SAIA_2025_Q3_SEC_ex99_1.htm",
        "publication_date": "2025-10-24",
        "section": "Financial highlights and operating statistics tables",
    },
    "SAIA_2025_Q4": {
        "url": "https://www.sec.gov/Archives/edgar/data/1177702/000119312526043714/saia-ex99_1.htm",
        "file": "SAIA_2025_Q4_SEC_ex99_1.htm",
        "publication_date": "2026-02-06",
        "section": "Financial highlights and operating statistics tables",
    },
    "SAIA_2026_Q1": {
        "url": "https://www.sec.gov/Archives/edgar/data/1177702/000119312526194062/saia-ex99_1.htm",
        "file": "SAIA_2026_Q1_SEC_ex99_1.htm",
        "publication_date": "2026-04-24",
        "section": "Financial highlights and operating statistics tables",
    },
    "XPO_2026_Q1": {
        "url": "https://www.sec.gov/Archives/edgar/data/1166003/000110465926052072/tm2612963d1_ex99-1.htm",
        "file": "XPO_2026_Q1_SEC_ex99_1.htm",
        "publication_date": "2026-04-30",
        "section": "North American LTL operating statistics and segment results",
    },
    "FDX_FY2026_Q1": {
        "url": "https://www.sec.gov/Archives/edgar/data/1048911/000104891125000042/fdx-earningsreleasefy2026q1.htm",
        "file": "FDX_FY2026_Q1_SEC_ex99_1.htm",
        "publication_date": "2025-09-18",
        "section": "FedEx Freight segment financial and operating highlights",
    },
    "FDX_FY2026_Q2": {
        "url": "https://www.sec.gov/Archives/edgar/data/1048911/000104891125000076/fdx-earningsreleasefy2026q2.htm",
        "file": "FDX_FY2026_Q2_SEC_ex99_1.htm",
        "publication_date": "2025-12-18",
        "section": "FedEx Freight segment financial and operating highlights",
    },
    "FDX_FY2026_Q3": {
        "url": "https://www.sec.gov/Archives/edgar/data/1048911/000104891126000010/fdx-earningsreleasefy2026q3.htm",
        "file": "FDX_FY2026_Q3_SEC_ex99_1.htm",
        "publication_date": "2026-03-19",
        "section": "FedEx Freight segment financial and operating highlights",
    },
    "FDXF_FY2026_Q4": {
        "url": "https://ir.fedexfreight.com/news-events/press-releases/detail/185/fedex-freight-reports-fourth-quarter-and-full-fiscal-year-2026-financial-results",
        "file": "FDXF_FY2026_Q4_IR.html",
        "publication_date": "2026-06-25",
        "section": "Fourth quarter fiscal 2026 results bullets and non-GAAP reconciliation",
    },
    "MASTIO_2025_ODFL": {
        "url": "https://ir.odfl.com/news-events/press-releases/detail/332/old-dominion-freight-line-selected-1-national-ltl-carrier",
        "file": "MASTIO_2025_ODFL_official_release.html",
        "publication_date": "2025-10-07",
        "section": "Mastio 2025 quality award official company release",
    },
    "MASTIO_2025_SECONDARY": {
        "url": "https://www.gain.consulting/post/2025-mastio-ltl-carrier-rankings-insights-into-customer-value-and-loyalty",
        "file": "MASTIO_2025_secondary_Gain_rankings.html",
        "publication_date": "2025-10-08",
        "section": "Public summary of 21st Edition national carrier rankings",
    },
    "FDX_STATBOOK_FY2026_Q3": {
        "url": "https://ae.marketscreener.com/news/fedex-third-quarter-2026-stat-book-ce7e5eded18cf12d",
        "file": "",
        "publication_date": "2026-03-19",
        "section": "FedEx Freight Segment - Quarterly Operating Statistics; FedEx Freight Segment - Quarterly Consolidated Income Statements",
    },
    "ODFL_2026_Q2": {
        "url": "https://www.sec.gov/Archives/edgar/data/878927/000087892726000021/odfl-ex99_1.htm",
        "file": "ODFL_2026_Q2_SEC_ex99_1.htm",
        "publication_date": "2026-07-29",
        "section": "Quarterly financial and operating statistics tables",
    },
    "XPO_2026_Q2": {
        "url": "https://www.sec.gov/Archives/edgar/data/1166003/000110465926088438/tm2616097d5_ex99-1.htm",
        "file": "XPO_2026_Q2_SEC_ex99_1.htm",
        "publication_date": "2026-07-30",
        "section": "North American LTL operating statistics and segment results",
    },
    "SAIA_2026_Q2": {
        "url": "https://www.sec.gov/Archives/edgar/data/1177702/000119312526324937/saia-ex99_1.htm",
        "file": "SAIA_2026_Q2_SEC_ex99_1.htm",
        "publication_date": "2026-07-30",
        "section": "Financial highlights and operating statistics tables",
    },
    "MASTIO_2021_ODFL": {
        "url": "https://ir.odfl.com/news-events/press-releases/detail/254/mastio-company-ranks-old-dominion-freight-line-no-1",
        "file": "",
        "publication_date": "2021-11-30",
        "section": "Official ODFL release quoting Mastio 2021 national LTL carrier quality award",
    },
    "MASTIO_2022_ODFL": {
        "url": "https://ir.odfl.com/news-events/press-releases/detail/273/old-dominion-freight-line-named-no-1-national-ltl-carrier",
        "file": "",
        "publication_date": "2022-10-11",
        "section": "Official ODFL release quoting Mastio 2022 national LTL carrier quality award",
    },
    "MASTIO_2023_ODFL": {
        "url": "https://www.businesswire.com/news/home/20231024169680/en/Old-Dominion-Freight-Line-Receives-Prestigious-Mastio-Quality-Award-for-the-14th-Consecutive-Year",
        "file": "",
        "publication_date": "2023-10-24",
        "section": "Official carrier release quoting Mastio 2023 national LTL carrier quality award",
    },
    "MASTIO_2023_TRADE": {
        "url": "https://www.freightwaves.com/news/ltl-survey-averitt-no-1-overall-old-dominion-top-national-carrier",
        "file": "",
        "publication_date": "2023-10-23",
        "section": "Trade publication reproducing 2023 Mastio national-carrier rankings",
    },
    "MASTIO_2024_TRADE": {
        "url": "https://www.freightwaves.com/news/ltl-survey-daylight-transport-no-1-overall-old-dominion-top-national-carrier",
        "file": "",
        "publication_date": "2024-10-15",
        "section": "Trade publication reproducing 2024 Mastio national-carrier rankings",
    },
}


RAW_HEADERS = [
    "observation_id", "carrier", "ticker", "period_start_date", "period_end_date",
    "calendar_year", "calendar_quarter", "fiscal_year_reported", "fiscal_quarter_reported",
    "comparison_period", "comparison_alignment", "operating_days", "reporting_basis",
    "revenue_m", "operating_income_m", "adjusted_operating_income_m",
    "operating_margin_pct", "adjusted_operating_margin_pct", "operating_ratio_pct",
    "adjusted_operating_ratio_pct", "avg_daily_shipments", "shipments_yoy_pct",
    "tonnage_per_day", "tonnage_yoy_pct", "weight_per_shipment_lbs", "weight_per_shipment_yoy_pct",
    "revenue_per_shipment", "revenue_per_shipment_yoy_pct", "revenue_per_shipment_ex_fuel",
    "revenue_per_shipment_ex_fuel_yoy_pct", "revenue_per_cwt", "revenue_per_cwt_yoy_pct",
    "revenue_per_cwt_ex_fuel", "revenue_per_cwt_ex_fuel_yoy_pct", "yield_reported",
    "yield_ex_fuel_reported", "base_revenue_per_cwt", "base_revenue_per_cwt_yoy_pct",
    "base_revenue_per_shipment", "base_revenue_per_shipment_yoy_pct",
    "fuel_surcharge_revenue_m", "fuel_surcharge_pct_revenue",
    "average_length_of_haul_miles", "load_factor", "comparability", "comparability_note",
    "source_key", "source_url", "local_source_file", "source_status", "extraction_status",
]


def row(source_key, carrier, ticker, start, end, cy, cq, fy, fq, comp, align, days, **metrics):
    src = SOURCES[source_key]
    local_file = str(SRC_DIR / src["file"]) if src.get("file") else ""
    source_status = "DOWNLOADED" if src.get("file") and (SRC_DIR / src["file"]).exists() else "URL_ONLY"
    base = {h: "" for h in RAW_HEADERS}
    base.update({
        "observation_id": f"{ticker}_{fy}_{fq}_{end}",
        "carrier": carrier,
        "ticker": ticker,
        "period_start_date": start,
        "period_end_date": end,
        "calendar_year": cy,
        "calendar_quarter": cq,
        "fiscal_year_reported": fy,
        "fiscal_quarter_reported": fq,
        "comparison_period": comp,
        "comparison_alignment": align,
        "operating_days": days,
        "reporting_basis": "Quarterly segment/company operating statistics",
        "source_key": source_key,
        "source_url": src["url"],
        "local_source_file": local_file,
        "source_status": source_status,
        "extraction_status": "EXTRACTED",
    })
    base.update(metrics)
    if base["operating_ratio_pct"] == "" and base["operating_margin_pct"] != "":
        base["operating_ratio_pct"] = round(100 - float(base["operating_margin_pct"]), 1)
    if base["operating_margin_pct"] == "" and base["operating_ratio_pct"] != "":
        base["operating_margin_pct"] = round(100 - float(base["operating_ratio_pct"]), 1)
    return base


rows = []

# Old Dominion Freight Line
rows += [
    row("ODFL_2024_Q1", "Old Dominion Freight Line", "ODFL", "2024-01-01", "2024-03-31", 2024, "Q1", "FY2024", "Q1", "Q1 2023", "EXACT", 64,
        revenue_m=1460.073, operating_income_m=386.426, operating_ratio_pct=73.5, avg_daily_shipments=46931,
        shipments_yoy_pct=-0.5, tonnage_per_day=35380, tonnage_yoy_pct=-3.2, weight_per_shipment_lbs=1508,
        weight_per_shipment_yoy_pct=-2.7, revenue_per_shipment=482.24, revenue_per_shipment_yoy_pct=1.3,
        revenue_per_shipment_ex_fuel=403.71, revenue_per_shipment_ex_fuel_yoy_pct=3.8,
        revenue_per_cwt=31.98, revenue_per_cwt_yoy_pct=4.1, revenue_per_cwt_ex_fuel=26.78,
        revenue_per_cwt_ex_fuel_yoy_pct=6.7, yield_ex_fuel_reported=6.7, comparability="Comparable",
        comparability_note="ODFL single LTL segment; company reports operating ratio and daily metrics."),
    row("ODFL_2024_Q2", "Old Dominion Freight Line", "ODFL", "2024-04-01", "2024-06-30", 2024, "Q2", "FY2024", "Q2", "Q2 2023", "EXACT", 64,
        revenue_m=1498.697, operating_income_m=421.691, operating_ratio_pct=71.9, avg_daily_shipments=48444,
        shipments_yoy_pct=3.1, tonnage_per_day=36560, tonnage_yoy_pct=1.9, weight_per_shipment_lbs=1509,
        weight_per_shipment_yoy_pct=-1.2, revenue_per_shipment=479.48, revenue_per_shipment_yoy_pct=3.2,
        revenue_per_shipment_ex_fuel=403.77, revenue_per_shipment_ex_fuel_yoy_pct=3.7,
        revenue_per_cwt=31.77, revenue_per_cwt_yoy_pct=4.4, revenue_per_cwt_ex_fuel=26.75,
        revenue_per_cwt_ex_fuel_yoy_pct=4.9, yield_ex_fuel_reported=4.9, comparability="Comparable",
        comparability_note="ODFL single LTL segment; company reports operating ratio and daily metrics."),
    row("ODFL_2024_Q3", "Old Dominion Freight Line", "ODFL", "2024-07-01", "2024-09-30", 2024, "Q3", "FY2024", "Q3", "Q3 2023", "EXACT", 64,
        revenue_m=1470.211, operating_income_m=401.861, operating_ratio_pct=72.7, avg_daily_shipments=47967,
        shipments_yoy_pct=-3.4, tonnage_per_day=35408, tonnage_yoy_pct=-4.8, weight_per_shipment_lbs=1476,
        weight_per_shipment_yoy_pct=-1.4, revenue_per_shipment=477.70, revenue_per_shipment_yoy_pct=0.1,
        revenue_per_shipment_ex_fuel=405.85, revenue_per_shipment_ex_fuel_yoy_pct=3.1,
        revenue_per_cwt=32.36, revenue_per_cwt_yoy_pct=1.5, revenue_per_cwt_ex_fuel=27.49,
        revenue_per_cwt_ex_fuel_yoy_pct=4.6, yield_ex_fuel_reported=4.6, comparability="Comparable",
        comparability_note="ODFL single LTL segment; company reports operating ratio and daily metrics."),
    row("ODFL_2024_Q4_PRIOR", "Old Dominion Freight Line", "ODFL", "2024-10-01", "2024-12-31", 2024, "Q4", "FY2024", "Q4", "Q4 2023", "NOT_ALIGNED", 62,
        revenue_m=1385.829, operating_income_m=334.020, operating_ratio_pct=75.9, avg_daily_shipments=45763,
        tonnage_per_day=34351, weight_per_shipment_lbs=1501, revenue_per_shipment=481.91,
        revenue_per_shipment_ex_fuel=413.21, revenue_per_cwt=32.10, revenue_per_cwt_ex_fuel=27.52,
        comparability="Comparable with caveat",
        comparability_note="Values sourced from prior-year column in 2025 Q4 comparative table; 2024-vs-2023 growth not in this source."),
    row("ODFL_2025_Q1", "Old Dominion Freight Line", "ODFL", "2025-01-01", "2025-03-31", 2025, "Q1", "FY2025", "Q1", "Q1 2024", "EXACT", 63,
        revenue_m=1374.858, operating_income_m=338.055, operating_ratio_pct=75.4, avg_daily_shipments=44566,
        shipments_yoy_pct=-5.0, tonnage_per_day=33135, tonnage_yoy_pct=-6.3, weight_per_shipment_lbs=1487,
        weight_per_shipment_yoy_pct=-1.4, revenue_per_shipment=485.79, revenue_per_shipment_yoy_pct=0.7,
        revenue_per_shipment_ex_fuel=414.68, revenue_per_shipment_ex_fuel_yoy_pct=2.7,
        revenue_per_cwt=32.67, revenue_per_cwt_yoy_pct=2.2, revenue_per_cwt_ex_fuel=27.89,
        revenue_per_cwt_ex_fuel_yoy_pct=4.1, yield_ex_fuel_reported=4.1, comparability="Comparable",
        comparability_note="ODFL single LTL segment; company reports operating ratio and daily metrics."),
    row("ODFL_2025_Q2", "Old Dominion Freight Line", "ODFL", "2025-04-01", "2025-06-30", 2025, "Q2", "FY2025", "Q2", "Q2 2024", "EXACT", 64,
        revenue_m=1407.724, operating_income_m=357.895, operating_ratio_pct=74.6, avg_daily_shipments=44907,
        shipments_yoy_pct=-7.3, tonnage_per_day=33178, tonnage_yoy_pct=-9.3, weight_per_shipment_lbs=1478,
        weight_per_shipment_yoy_pct=-2.1, revenue_per_shipment=485.31, revenue_per_shipment_yoy_pct=1.2,
        revenue_per_shipment_ex_fuel=416.31, revenue_per_shipment_ex_fuel_yoy_pct=3.1,
        revenue_per_cwt=32.84, revenue_per_cwt_yoy_pct=3.4, revenue_per_cwt_ex_fuel=28.17,
        revenue_per_cwt_ex_fuel_yoy_pct=5.3, yield_ex_fuel_reported=5.3, comparability="Comparable",
        comparability_note="ODFL single LTL segment; company reports operating ratio and daily metrics."),
    row("ODFL_2025_Q3", "Old Dominion Freight Line", "ODFL", "2025-07-01", "2025-09-30", 2025, "Q3", "FY2025", "Q3", "Q3 2024", "EXACT", 64,
        revenue_m=1406.511, operating_income_m=360.844, operating_ratio_pct=74.3, avg_daily_shipments=44201,
        shipments_yoy_pct=-7.9, tonnage_per_day=32231, tonnage_yoy_pct=-9.0, weight_per_shipment_lbs=1458,
        weight_per_shipment_yoy_pct=-1.2, revenue_per_shipment=494.17, revenue_per_shipment_yoy_pct=3.4,
        revenue_per_shipment_ex_fuel=419.67, revenue_per_shipment_ex_fuel_yoy_pct=3.4,
        revenue_per_cwt=33.88, revenue_per_cwt_yoy_pct=4.7, revenue_per_cwt_ex_fuel=28.78,
        revenue_per_cwt_ex_fuel_yoy_pct=4.7, yield_ex_fuel_reported=4.7, comparability="Comparable",
        comparability_note="ODFL single LTL segment; company reports operating ratio and daily metrics."),
    row("ODFL_2025_Q4", "Old Dominion Freight Line", "ODFL", "2025-10-01", "2025-12-31", 2025, "Q4", "FY2025", "Q4", "Q4 2024", "EXACT", 62,
        revenue_m=1307.296, operating_income_m=304.251, operating_ratio_pct=76.7, avg_daily_shipments=41308,
        shipments_yoy_pct=-9.7, tonnage_per_day=30691, tonnage_yoy_pct=-10.7, weight_per_shipment_lbs=1486,
        weight_per_shipment_yoy_pct=-1.0, revenue_per_shipment=503.95, revenue_per_shipment_yoy_pct=4.6,
        revenue_per_shipment_ex_fuel=429.21, revenue_per_shipment_ex_fuel_yoy_pct=3.9,
        revenue_per_cwt=33.91, revenue_per_cwt_yoy_pct=5.6, revenue_per_cwt_ex_fuel=28.88,
        revenue_per_cwt_ex_fuel_yoy_pct=4.9, yield_ex_fuel_reported=4.9, comparability="Comparable",
        comparability_note="ODFL single LTL segment; company reports operating ratio and daily metrics."),
    row("ODFL_2026_Q1", "Old Dominion Freight Line", "ODFL", "2026-01-01", "2026-03-31", 2026, "Q1", "FY2026", "Q1", "Q1 2025", "EXACT", 63,
        revenue_m=1334.696, operating_income_m=317.341, operating_ratio_pct=76.2, avg_daily_shipments=41037,
        shipments_yoy_pct=-7.9, tonnage_per_day=30584, tonnage_yoy_pct=-7.7, weight_per_shipment_lbs=1491,
        weight_per_shipment_yoy_pct=0.3, revenue_per_shipment=514.56, revenue_per_shipment_yoy_pct=5.9,
        revenue_per_shipment_ex_fuel=434.16, revenue_per_shipment_ex_fuel_yoy_pct=4.7,
        revenue_per_cwt=34.52, revenue_per_cwt_yoy_pct=5.7, revenue_per_cwt_ex_fuel=29.13,
        revenue_per_cwt_ex_fuel_yoy_pct=4.4, yield_ex_fuel_reported=4.4, comparability="Comparable",
        comparability_note="ODFL single LTL segment; company reports operating ratio and daily metrics."),
]

# Saia
rows += [
    row("SAIA_2024_Q1", "Saia", "SAIA", "2024-01-01", "2024-03-31", 2024, "Q1", "FY2024", "Q1", "Q1 2023", "EXACT", 64,
        revenue_m=754.775, operating_income_m=117.912, operating_ratio_pct=84.4, avg_daily_shipments=32940,
        shipments_yoy_pct=15.7, tonnage_per_day=21750, tonnage_yoy_pct=6.2, weight_per_shipment_lbs=1321,
        weight_per_shipment_yoy_pct=-8.2, revenue_per_shipment=350.18, revenue_per_shipment_yoy_pct=-1.2,
        revenue_per_shipment_ex_fuel=293.96, revenue_per_shipment_ex_fuel_yoy_pct=1.4,
        revenue_per_cwt=26.51, revenue_per_cwt_yoy_pct=7.6, revenue_per_cwt_ex_fuel=22.26,
        revenue_per_cwt_ex_fuel_yoy_pct=10.5, yield_ex_fuel_reported=10.5, comparability="Comparable with caveat",
        comparability_note="Saia tonnage and shipment rows are reported in thousands; daily values multiplied by 1,000."),
    row("SAIA_2024_Q2", "Saia", "SAIA", "2024-04-01", "2024-06-30", 2024, "Q2", "FY2024", "Q2", "Q2 2023", "EXACT", 64,
        revenue_m=823.244, operating_income_m=137.593, operating_ratio_pct=83.3, avg_daily_shipments=36360,
        shipments_yoy_pct=18.1, tonnage_per_day=24360, tonnage_yoy_pct=9.7, weight_per_shipment_lbs=1340,
        weight_per_shipment_yoy_pct=-7.1, revenue_per_shipment=345.07, revenue_per_shipment_yoy_pct=0.3,
        revenue_per_shipment_ex_fuel=290.72, revenue_per_shipment_ex_fuel_yoy_pct=1.0,
        revenue_per_cwt=25.75, revenue_per_cwt_yoy_pct=8.0, revenue_per_cwt_ex_fuel=21.69,
        revenue_per_cwt_ex_fuel_yoy_pct=8.7, yield_ex_fuel_reported=8.7, comparability="Comparable with caveat",
        comparability_note="Saia tonnage and shipment rows are reported in thousands; daily values multiplied by 1,000."),
    row("SAIA_2024_Q3", "Saia", "SAIA", "2024-07-01", "2024-09-30", 2024, "Q3", "FY2024", "Q3", "Q3 2023", "APPROXIMATE", 64,
        revenue_m=842.103, operating_income_m=125.171, operating_ratio_pct=85.1, avg_daily_shipments=37170,
        shipments_yoy_pct=8.5, tonnage_per_day=25080, tonnage_yoy_pct=7.7, weight_per_shipment_lbs=1349,
        weight_per_shipment_yoy_pct=-0.8, revenue_per_shipment=345.93, revenue_per_shipment_yoy_pct=-1.6,
        revenue_per_shipment_ex_fuel=293.39, revenue_per_shipment_ex_fuel_yoy_pct=0.9,
        revenue_per_cwt=25.64, revenue_per_cwt_yoy_pct=-0.9, revenue_per_cwt_ex_fuel=21.75,
        revenue_per_cwt_ex_fuel_yoy_pct=1.7, yield_ex_fuel_reported=1.7, comparability="Comparable with caveat",
        comparability_note="Prior-year quarter had 63 workdays; source reports daily-growth comparison."),
    row("SAIA_2024_Q4", "Saia", "SAIA", "2024-10-01", "2024-12-31", 2024, "Q4", "FY2024", "Q4", "Q4 2023", "APPROXIMATE", 62,
        revenue_m=788.952, operating_income_m=101.484, operating_ratio_pct=87.1, avg_daily_shipments=35060,
        shipments_yoy_pct=4.5, tonnage_per_day=23890, tonnage_yoy_pct=8.3, weight_per_shipment_lbs=1362,
        weight_per_shipment_yoy_pct=3.7, revenue_per_shipment=350.51, revenue_per_shipment_yoy_pct=-2.0,
        revenue_per_shipment_ex_fuel=299.17, revenue_per_shipment_ex_fuel_yoy_pct=1.3,
        revenue_per_cwt=25.73, revenue_per_cwt_yoy_pct=-5.4, revenue_per_cwt_ex_fuel=21.96,
        revenue_per_cwt_ex_fuel_yoy_pct=-2.3, yield_ex_fuel_reported=-2.3, comparability="Comparable with caveat",
        comparability_note="Prior-year quarter had 61 workdays; source reports daily-growth comparison."),
    row("SAIA_2025_Q1", "Saia", "SAIA", "2025-01-01", "2025-03-31", 2025, "Q1", "FY2025", "Q1", "Q1 2024", "APPROXIMATE", 63,
        revenue_m=787.575, operating_income_m=70.168, operating_ratio_pct=91.1, avg_daily_shipments=34440,
        shipments_yoy_pct=4.6, tonnage_per_day=24520, tonnage_yoy_pct=12.7, weight_per_shipment_lbs=1424,
        weight_per_shipment_yoy_pct=7.8, revenue_per_shipment=355.48, revenue_per_shipment_yoy_pct=1.5,
        revenue_per_shipment_ex_fuel=300.76, revenue_per_shipment_ex_fuel_yoy_pct=2.3,
        revenue_per_cwt=24.97, revenue_per_cwt_yoy_pct=-5.8, revenue_per_cwt_ex_fuel=21.12,
        revenue_per_cwt_ex_fuel_yoy_pct=-5.1, yield_ex_fuel_reported=-5.1, comparability="Comparable with caveat",
        comparability_note="Prior-year quarter had 64 workdays; source reports daily-growth comparison."),
    row("SAIA_2025_Q2", "Saia", "SAIA", "2025-04-01", "2025-06-30", 2025, "Q2", "FY2025", "Q2", "Q2 2024", "EXACT", 64,
        revenue_m=817.115, operating_income_m=99.399, operating_ratio_pct=87.8, avg_daily_shipments=35330,
        shipments_yoy_pct=-2.8, tonnage_per_day=24630, tonnage_yoy_pct=1.1, weight_per_shipment_lbs=1394,
        weight_per_shipment_yoy_pct=4.0, revenue_per_shipment=351.36, revenue_per_shipment_yoy_pct=1.8,
        revenue_per_shipment_ex_fuel=298.71, revenue_per_shipment_ex_fuel_yoy_pct=2.7,
        revenue_per_cwt=25.20, revenue_per_cwt_yoy_pct=-2.1, revenue_per_cwt_ex_fuel=21.42,
        revenue_per_cwt_ex_fuel_yoy_pct=-1.2, yield_ex_fuel_reported=-1.2, comparability="Comparable with caveat",
        comparability_note="Saia tonnage and shipment rows are reported in thousands; daily values multiplied by 1,000."),
    row("SAIA_2025_Q3", "Saia", "SAIA", "2025-07-01", "2025-09-30", 2025, "Q3", "FY2025", "Q3", "Q3 2024", "EXACT", 64,
        revenue_m=839.644, operating_income_m=118.610, operating_ratio_pct=85.9, avg_daily_shipments=36450,
        shipments_yoy_pct=-1.9, tonnage_per_day=24700, tonnage_yoy_pct=-1.5, weight_per_shipment_lbs=1355,
        weight_per_shipment_yoy_pct=0.4, revenue_per_shipment=349.07, revenue_per_shipment_yoy_pct=0.9,
        revenue_per_shipment_ex_fuel=294.35, revenue_per_shipment_ex_fuel_yoy_pct=0.3,
        revenue_per_cwt=25.76, revenue_per_cwt_yoy_pct=0.5, revenue_per_cwt_ex_fuel=21.72,
        revenue_per_cwt_ex_fuel_yoy_pct=-0.1, yield_ex_fuel_reported=-0.1, comparability="Comparable with caveat",
        comparability_note="Saia tonnage and shipment rows are reported in thousands; daily values multiplied by 1,000."),
    row("SAIA_2025_Q4", "Saia", "SAIA", "2025-10-01", "2025-12-31", 2025, "Q4", "FY2025", "Q4", "Q4 2024", "EXACT", 62,
        revenue_m=789.952, operating_income_m=64.023, operating_ratio_pct=91.9, avg_daily_shipments=34900,
        shipments_yoy_pct=-0.5, tonnage_per_day=23530, tonnage_yoy_pct=-1.5, weight_per_shipment_lbs=1348,
        weight_per_shipment_yoy_pct=-1.0, revenue_per_shipment=352.27, revenue_per_shipment_yoy_pct=0.5,
        revenue_per_shipment_ex_fuel=297.57, revenue_per_shipment_ex_fuel_yoy_pct=-0.5,
        revenue_per_cwt=26.13, revenue_per_cwt_yoy_pct=1.6, revenue_per_cwt_ex_fuel=22.07,
        revenue_per_cwt_ex_fuel_yoy_pct=0.5, yield_ex_fuel_reported=0.5, comparability="Comparable with caveat",
        comparability_note="Saia tonnage and shipment rows are reported in thousands; daily values multiplied by 1,000."),
    row("SAIA_2026_Q1", "Saia", "SAIA", "2026-01-01", "2026-03-31", 2026, "Q1", "FY2026", "Q1", "Q1 2025", "EXACT", 63,
        revenue_m=806.226, operating_income_m=66.806, operating_ratio_pct=91.7, avg_daily_shipments=34790,
        shipments_yoy_pct=1.0, tonnage_per_day=24020, tonnage_yoy_pct=-2.1, weight_per_shipment_lbs=1380,
        weight_per_shipment_yoy_pct=-3.1, revenue_per_shipment=357.93, revenue_per_shipment_yoy_pct=0.7,
        revenue_per_shipment_ex_fuel=297.11, revenue_per_shipment_ex_fuel_yoy_pct=-1.2,
        revenue_per_cwt=25.93, revenue_per_cwt_yoy_pct=3.8, revenue_per_cwt_ex_fuel=21.52,
        revenue_per_cwt_ex_fuel_yoy_pct=1.9, yield_ex_fuel_reported=1.9, comparability="Comparable with caveat",
        comparability_note="Saia tonnage and shipment rows are reported in thousands; daily values multiplied by 1,000."),
]

# XPO North American LTL. XPO reports pounds/day in thousands; tonnage_per_day is converted to tons/day.
xpo_specs = [
    ("2024_Q1", "2024-01-01", "2024-03-31", 2024, "Q1", 63.5, 1221, 165, 175, 85.7, 51392, 4.7, 70709, 2.6, 1376, 373.88, 309.57, 27.80, 23.13, 9.8, 848.3, 22869, "2024-05-03", "https://investors.xpo.com/news-releases/news-release-details/xpo-reports-first-quarter-2024-results"),
    ("2024_Q2", "2024-04-01", "2024-06-30", 2024, "Q2", 64.0, 1272, 203, 214, 83.2, 53519, 4.5, 72658, 3.4, 1358, 370.98, 310.24, 28.04, 23.56, 9.0, 847.8, 22884, "2024-08-01", "https://investors.xpo.com/news-releases/news-release-details/xpo-reports-second-quarter-2024-results/"),
    ("2024_Q3", "2024-07-01", "2024-09-30", 2024, "Q3", 63.5, 1251, 188, 198, 84.2, 51921, -3.2, 69470, -3.9, 1338, 379.00, 319.75, 28.77, 24.34, 6.7, 855.7, 22644, "2024-10-30", "https://investors.xpo.com/news-releases/news-release-details/xpo-reports-third-quarter-2024-results"),
    ("2024_Q4", "2024-10-01", "2024-12-31", 2024, "Q4", 61.5, 1156, 179, 159, 86.2, 49109, -4.4, 65433, -5.7, 1332, 382.32, 325.62, 29.09, 24.84, 6.3, 854.7, 22128, "2025-02-06", "https://investors.xpo.com/news-releases/news-release-details/xpo-reports-fourth-quarter-and-full-year-2024-results"),
    ("2025_Q1", "2025-01-01", "2025-03-31", 2025, "Q1", 63.0, 1172, 158, 165, 85.9, 48400, -5.8, 65427, -7.5, 1352, 384.27, 325.74, 29.06, 24.73, 6.9, 845.6, 22434, "2025-05-01", "https://investors.xpo.com/news-releases/news-release-details/xpo-reports-first-quarter-2025-results"),
    ("2025_Q2", "2025-04-01", "2025-06-30", 2025, "Q2", 63.5, 1240, 199, 211, 82.9, 50782, -5.1, 67813, -6.7, 1335, 384.13, 327.53, 29.23, 24.99, 6.1, 845.5, 22765, "2025-07-31", "https://investors.xpo.com/news-releases/news-release-details/xpo-reports-second-quarter-2025-results/"),
    ("2025_Q3", "2025-07-01", "2025-09-30", 2025, "Q3", 64.0, 1255, 208, 217, 82.7, 50094, -3.5, 65236, -6.1, 1302, 391.13, 330.48, 30.42, 25.77, 5.9, 866.7, 22442, "2025-10-30", "https://investors.xpo.com/news-releases/news-release-details/xpo-reports-third-quarter-2025-results"),
    ("2025_Q4", "2025-10-01", "2025-12-31", 2025, "Q4", 61.0, 1165, 184, 181, 84.4, 48348, -1.6, 62486, -4.5, 1292, 394.78, 335.28, 30.72, 26.12, 5.2, 862.8, 22002, "2026-02-05", "https://investors.xpo.com/news-releases/news-release-details/xpo-reports-fourth-quarter-2025-results"),
    ("2026_Q1", "2026-01-01", "2026-03-31", 2026, "Q1", 62.5, 1229, 189, 198, 83.9, 49834, 3.0, 65510, 0.1, 1315, 394.14, 329.77, 30.61, 25.71, 4.0, 852.6, 22294, "2026-04-30", SOURCES["XPO_2026_Q1"]["url"]),
]
for suffix, start, end, cy, cq, days, rev, oi, adj_oi, adj_or, shipments, ship_yoy, pounds_k_day, ton_yoy, wt, rps, rps_ex, rcwt, rcwt_ex, yield_ex, haul, load, pub_date, url in xpo_specs:
    src_key = "XPO_2026_Q1" if suffix == "2026_Q1" else f"XPO_URL_ONLY_{suffix}"
    if src_key not in SOURCES:
        SOURCES[src_key] = {
            "url": url,
            "file": "",
            "publication_date": pub_date,
            "section": "North American LTL operating statistics and segment results",
        }
    rows.append(row(src_key, "XPO North American LTL", "XPO", start, end, cy, cq, f"FY{cy}", cq, f"{cq} {cy - 1}", "EXACT", days,
        revenue_m=rev, operating_income_m=oi, adjusted_operating_income_m=adj_oi,
        operating_ratio_pct=round(100 - (oi / rev * 100), 1), adjusted_operating_ratio_pct=adj_or,
        adjusted_operating_margin_pct=round(100 - adj_or, 1), avg_daily_shipments=shipments,
        shipments_yoy_pct=ship_yoy, tonnage_per_day=round(pounds_k_day * 1000 / 2000, 1), tonnage_yoy_pct=ton_yoy,
        weight_per_shipment_lbs=wt, revenue_per_shipment=rps, revenue_per_shipment_ex_fuel=rps_ex,
        revenue_per_cwt=rcwt, revenue_per_cwt_ex_fuel=rcwt_ex, revenue_per_cwt_ex_fuel_yoy_pct=yield_ex,
        yield_ex_fuel_reported=yield_ex, average_length_of_haul_miles=haul, load_factor=load, comparability="Comparable with caveat",
        comparability_note="XPO source reports pounds/day in thousands; normalized tonnage/day converts pounds to tons. XPO 2024-2025 source pages retained as URL-only after local downloads timed out."))

# FedEx Freight
rows += [
    row("FDX_FY2026_Q1", "FedEx Freight", "FDXF", "2025-06-01", "2025-08-31", 2025, "Q3", "FY2026", "Q1", "FY2025 Q1", "APPROXIMATE", 64,
        revenue_m=2257, operating_income_m=360, operating_margin_pct=16.0, avg_daily_shipments=90000,
        shipments_yoy_pct=-2, weight_per_shipment_lbs=925, revenue_per_shipment=374.62, revenue_per_shipment_yoy_pct=-1,
        revenue_per_cwt=40.50, revenue_per_cwt_yoy_pct=-1, comparability="Comparable with caveat",
        comparability_note="FedEx reports fiscal quarters and composite revenue metrics; not calendar-quarter aligned with peers."),
    row("FDX_FY2026_Q2", "FedEx Freight", "FDXF", "2025-09-01", "2025-11-30", 2025, "Q4", "FY2026", "Q2", "FY2025 Q2", "APPROXIMATE", 62,
        revenue_m=2139, operating_income_m=90, operating_margin_pct=4.2, avg_daily_shipments=87400,
        shipments_yoy_pct=-4, weight_per_shipment_lbs=924, weight_per_shipment_yoy_pct=1,
        revenue_per_shipment=375.97, revenue_per_shipment_yoy_pct=2, revenue_per_cwt=40.71,
        revenue_per_cwt_yoy_pct=1, comparability="Comparable with caveat",
        comparability_note="FedEx reports fiscal quarters and composite revenue metrics; not calendar-quarter aligned with peers."),
    row("FDX_FY2026_Q3", "FedEx Freight", "FDXF", "2025-12-01", "2026-02-28", 2026, "Q1", "FY2026", "Q3", "FY2025 Q3", "APPROXIMATE", "",
        revenue_m=1991, operating_income_m=8, operating_margin_pct=0.4, avg_daily_shipments=80200,
        shipments_yoy_pct=-6, weight_per_shipment_lbs=926, weight_per_shipment_yoy_pct=1,
        revenue_per_shipment=380.24, revenue_per_shipment_yoy_pct=1, revenue_per_cwt=41.08,
        revenue_per_cwt_yoy_pct=0, comparability="Comparable with caveat",
        comparability_note="FedEx reports fiscal quarters and composite revenue metrics; not calendar-quarter aligned with peers."),
    row("FDXF_FY2026_Q4", "FedEx Freight", "FDXF", "2026-03-01", "2026-05-31", 2026, "Q2", "FY2026", "Q4", "FY2025 Q4", "APPROXIMATE", "",
        revenue_m=2400, operating_income_m=158, adjusted_operating_income_m=363, operating_margin_pct=6.6,
        adjusted_operating_margin_pct=15.1, avg_daily_shipments=86700, shipments_yoy_pct=-5.9,
        weight_per_shipment_lbs=948, weight_per_shipment_yoy_pct=3.0, revenue_per_shipment=415.22,
        revenue_per_shipment_yoy_pct=11.5, revenue_per_cwt=43.79, revenue_per_cwt_yoy_pct=8.2,
        base_revenue_per_cwt_yoy_pct="slight decline", comparability="Comparable with caveat",
        comparability_note="FedEx reports fiscal quarters and composite metrics; Q4 source reports base revenue/CWT only directionally.")
]

fedex_history_specs = [
    ("FY2024", "Q1", "2023-06-01", "2023-08-31", 2023, "Q3", 65, 2385, 482, 20.2, 94635, -12.6, 955, -6.4, 369.56, -4.2, 38.71, 2.4),
    ("FY2024", "Q2", "2023-09-01", "2023-11-30", 2023, "Q4", 62, 2452, 491, 20.0, 99001, -5.1, 946, -5.6, 381.05, 0.9, 40.29, 7.0),
    ("FY2024", "Q3", "2023-12-01", "2024-02-29", 2024, "Q1", 62, 2205, 341, 15.5, 89248, -4.1, 946, -3.2, 379.26, -0.7, 40.10, 2.6),
    ("FY2024", "Q4", "2024-03-01", "2024-05-31", 2024, "Q2", 65, 2387, 507, 21.2, 93080, 0.1, 939, -2.8, 377.63, 1.1, 40.22, 4.0),
    ("FY2025", "Q1", "2024-06-01", "2024-08-31", 2024, "Q3", 64, 2329, 439, 18.8, 92008, -2.8, 928, -2.8, 378.09, 2.3, 40.73, 5.2),
    ("FY2025", "Q2", "2024-09-01", "2024-11-30", 2024, "Q4", 62, 2177, 312, 14.3, 90998, -8.1, 913, -3.5, 367.60, -3.5, 40.26, -0.1),
    ("FY2025", "Q3", "2024-12-01", "2025-02-28", 2025, "Q1", 62, 2089, 261, 12.5, 85072, -4.7, 917, -3.1, 375.81, -0.9, 41.00, 2.2),
    ("FY2025", "Q4", "2025-03-01", "2025-05-31", 2025, "Q2", 64, 2297, 477, 20.8, 92129, -1.0, 920, -2.0, 372.55, -1.3, 40.49, 0.7),
]
for fy, fq, start, end, cy, cq, days, rev, oi, margin, shipments, ship_yoy, weight, weight_yoy, rps, rps_yoy, rcwt, rcwt_yoy in fedex_history_specs:
    rows.append(row("FDX_STATBOOK_FY2026_Q3", "FedEx Freight", "FDXF", start, end, cy, cq, fy, fq, fy.replace("FY", "FY") + " prior-year quarter", "APPROXIMATE", days,
        revenue_m=rev, operating_income_m=oi, operating_margin_pct=margin, avg_daily_shipments=shipments,
        shipments_yoy_pct=ship_yoy, weight_per_shipment_lbs=weight, weight_per_shipment_yoy_pct=weight_yoy,
        revenue_per_shipment=rps, revenue_per_shipment_yoy_pct=rps_yoy, revenue_per_cwt=rcwt,
        revenue_per_cwt_yoy_pct=rcwt_yoy, comparability="Comparable with caveat",
        comparability_note="FedEx historical statistical book. FY2025 values include FedEx Custom Critical in the Freight segment; prior-period figures are subsequently recast where shown by FedEx."))

for r in rows:
    if r["ticker"] == "FDXF" and r["fiscal_year_reported"] == "FY2026":
        adjusted_margin = {"Q1": 16.3, "Q2": 11.3, "Q3": 6.7, "Q4": 15.1}.get(r["fiscal_quarter_reported"])
        adjusted_oi = {"Q1": 369, "Q2": 242, "Q3": 134, "Q4": 363}.get(r["fiscal_quarter_reported"])
        if adjusted_margin is not None:
            r["adjusted_operating_margin_pct"] = adjusted_margin
            r["adjusted_operating_ratio_pct"] = round(100 - adjusted_margin, 1)
            r["adjusted_operating_income_m"] = adjusted_oi
            r["comparability_note"] += " Adjusted operating ratio is derived as 100 minus adjusted operating margin; adjusted metrics are the preferred profitability comparison during spin-off expense periods."
    if r["ticker"] == "FDXF" and r["fiscal_year_reported"] == "FY2026" and r["fiscal_quarter_reported"] in ("Q1", "Q2", "Q3"):
        exact_shipments = {"Q1": 90012, "Q2": 87438, "Q3": 80233}[r["fiscal_quarter_reported"]]
        r["avg_daily_shipments"] = exact_shipments
        if r["fiscal_quarter_reported"] == "Q3":
            r["operating_days"] = 62

peer_q2_rows = [
    row("ODFL_2026_Q2", "Old Dominion Freight Line", "ODFL", "2026-04-01", "2026-06-30", 2026, "Q2", "FY2026", "Q2", "Q2 2025", "EXACT", 64,
        revenue_m=1554.004, operating_income_m=465.301, operating_ratio_pct=70.1, avg_daily_shipments=42332,
        shipments_yoy_pct=-5.7, tonnage_per_day=31804, tonnage_yoy_pct=-4.1, weight_per_shipment_lbs=1503,
        weight_per_shipment_yoy_pct=1.7, revenue_per_shipment=568.55, revenue_per_shipment_yoy_pct=17.2,
        revenue_per_shipment_ex_fuel=446.44, revenue_per_shipment_ex_fuel_yoy_pct=7.2, revenue_per_cwt=37.84,
        revenue_per_cwt_yoy_pct=15.2, revenue_per_cwt_ex_fuel=29.71, revenue_per_cwt_ex_fuel_yoy_pct=5.5,
        yield_ex_fuel_reported=5.5, comparability="Comparable",
        comparability_note="Q2 2026 primary SEC exhibit; source also disclosed on-time service above 99% and cargo claims ratio 0.1%."),
    row("XPO_2026_Q2", "XPO North American LTL", "XPO", "2026-04-01", "2026-06-30", 2026, "Q2", "FY2026", "Q2", "Q2 2025", "EXACT", 63.5,
        revenue_m=1428, operating_income_m=285, adjusted_operating_income_m=287, operating_margin_pct=20.0,
        operating_ratio_pct=80.0, adjusted_operating_margin_pct=20.1, adjusted_operating_ratio_pct=79.9,
        avg_daily_shipments=52229, shipments_yoy_pct=2.8, tonnage_per_day=34231.5, tonnage_yoy_pct=1.0,
        weight_per_shipment_lbs=1311, weight_per_shipment_yoy_pct=-1.8, revenue_per_shipment=429.98,
        revenue_per_shipment_yoy_pct=11.9, revenue_per_shipment_ex_fuel=335.27,
        revenue_per_shipment_ex_fuel_yoy_pct=2.4, revenue_per_cwt=33.32, revenue_per_cwt_yoy_pct=14.0,
        revenue_per_cwt_ex_fuel=26.09, revenue_per_cwt_ex_fuel_yoy_pct=4.4, yield_ex_fuel_reported=4.4,
        average_length_of_haul_miles=853.6, load_factor=22287, comparability="Comparable with caveat",
        comparability_note="Q2 2026 primary SEC exhibit. XPO source reports pounds/day in thousands; normalized tonnage/day converts pounds to tons. Damage claims ratio was less than 0.2%."),
    row("SAIA_2026_Q2", "Saia", "SAIA", "2026-04-01", "2026-06-30", 2026, "Q2", "FY2026", "Q2", "Q2 2025", "EXACT", 64,
        revenue_m=956.494, operating_income_m=125.212, operating_ratio_pct=86.9, avg_daily_shipments=36890,
        shipments_yoy_pct=4.4, tonnage_per_day=26700, tonnage_yoy_pct=8.4, weight_per_shipment_lbs=1448,
        weight_per_shipment_yoy_pct=3.9, revenue_per_shipment=393.56, revenue_per_shipment_yoy_pct=12.0,
        revenue_per_shipment_ex_fuel=303.12, revenue_per_shipment_ex_fuel_yoy_pct=1.5, revenue_per_cwt=27.18,
        revenue_per_cwt_yoy_pct=7.9, revenue_per_cwt_ex_fuel=20.94, revenue_per_cwt_ex_fuel_yoy_pct=-2.2,
        yield_ex_fuel_reported=-2.2, average_length_of_haul_miles=888, comparability="Comparable with caveat",
        comparability_note="Q2 2026 primary SEC exhibit; Saia tonnage and shipment rows are reported in thousands and multiplied by 1,000. Source also disclosed claims ratio of 0.3%."),
]
rows.extend(peer_q2_rows)


CORE_FIELDS = [
    "revenue_m", "operating_income_m", "adjusted_operating_income_m", "operating_margin_pct",
    "adjusted_operating_margin_pct", "operating_ratio_pct", "adjusted_operating_ratio_pct",
    "avg_daily_shipments", "shipments_yoy_pct", "tonnage_per_day", "tonnage_yoy_pct",
    "weight_per_shipment_lbs", "revenue_per_shipment", "revenue_per_shipment_ex_fuel",
    "revenue_per_cwt", "revenue_per_cwt_ex_fuel", "revenue_per_cwt_ex_fuel_yoy_pct",
]


def metric_labels():
    return {
        "revenue_m": "Revenue / Operating Revenue / Segment revenue",
        "operating_income_m": "Operating income / Operating Income",
        "adjusted_operating_income_m": "Adjusted operating income / Non-GAAP measure",
        "operating_margin_pct": "Operating margin / derived from operating ratio where applicable",
        "adjusted_operating_margin_pct": "Adjusted operating margin / Non-GAAP measure",
        "operating_ratio_pct": "Operating ratio / derived from operating margin where applicable",
        "adjusted_operating_ratio_pct": "Adjusted operating ratio",
        "avg_daily_shipments": "LTL shipments per day / Total average daily shipments / Shipments per day",
        "shipments_yoy_pct": "LTL shipments per day change / Total average daily shipments change",
        "tonnage_per_day": "LTL tonnage per day / pounds per day converted to tons",
        "tonnage_yoy_pct": "LTL tonnage per day change / pounds per day change",
        "weight_per_shipment_lbs": "LTL weight per shipment (lbs.) / Composite weight per shipment",
        "weight_per_shipment_yoy_pct": "Weight per shipment year-over-year change",
        "revenue_per_shipment": "LTL revenue per shipment / Composite revenue per shipment",
        "revenue_per_shipment_yoy_pct": "Revenue per shipment year-over-year change",
        "revenue_per_shipment_ex_fuel": "LTL revenue per shipment excluding fuel surcharge",
        "revenue_per_shipment_ex_fuel_yoy_pct": "Revenue per shipment excluding fuel surcharge year-over-year change",
        "revenue_per_cwt": "LTL revenue per hundredweight / Composite revenue per hundredweight",
        "revenue_per_cwt_yoy_pct": "Revenue per hundredweight year-over-year change",
        "revenue_per_cwt_ex_fuel": "LTL revenue per hundredweight excluding fuel surcharge",
        "revenue_per_cwt_ex_fuel_yoy_pct": "Revenue per hundredweight excluding fuel surcharge year-over-year change",
        "yield_ex_fuel_reported": "Yield excluding fuel surcharge growth where source uses yield language",
        "base_revenue_per_cwt_yoy_pct": "Base revenue per hundredweight year-over-year change",
    }


def build_provenance():
    labels = metric_labels()
    prov = []
    for r in rows:
        src = SOURCES[r["source_key"]]
        for field, label in labels.items():
            if r.get(field) not in ("", None):
                method = "Reported"
                if field == "tonnage_per_day" and r["ticker"] == "XPO":
                    method = "Derived: reported pounds/day in thousands * 1000 / 2000"
                if field in ("operating_margin_pct", "operating_ratio_pct") and "derived" in label:
                    method = "Reported or algebraic complement when only reciprocal ratio is reported"
                if field == "avg_daily_shipments" and r["ticker"] in ("SAIA", "FDXF"):
                    method = "Reported in thousands; multiplied by 1,000 for count scale"
                prov.append({
                    "observation_id": r["observation_id"],
                    "carrier": r["carrier"],
                    "period_end_date": r["period_end_date"],
                    "field_name": field,
                    "reported_or_normalized_value": r[field],
                    "exact_reported_metric_label": label,
                    "source_publication_date": src.get("publication_date", ""),
                    "source_page_or_section": src.get("section", ""),
                    "source_url": r["source_url"],
                    "local_source_file": r["local_source_file"],
                    "extraction_method": method,
                    "confidence": "High" if r["source_status"] == "DOWNLOADED" else "Medium",
                    "notes": r["comparability_note"],
                })
    return prov


def completeness_matrix():
    expected = []
    calendar_periods = [
        ("2024-03-31", "2024 Q1"), ("2024-06-30", "2024 Q2"), ("2024-09-30", "2024 Q3"),
        ("2024-12-31", "2024 Q4"), ("2025-03-31", "2025 Q1"), ("2025-06-30", "2025 Q2"),
        ("2025-09-30", "2025 Q3"), ("2025-12-31", "2025 Q4"), ("2026-03-31", "2026 Q1"),
        ("2026-06-30", "2026 Q2"),
    ]
    for ticker, carrier in [("ODFL", "Old Dominion Freight Line"), ("XPO", "XPO North American LTL"), ("SAIA", "Saia")]:
        for end, label in calendar_periods:
            found = next((r for r in rows if r["ticker"] == ticker and r["period_end_date"] == end), None)
            expected.append({
                "carrier": carrier, "ticker": ticker, "period_label": label, "period_end_date": end,
                "expected": "Y", "source_found": "Y" if found else "N", "downloaded": found["source_status"] if found else "N",
                "raw_populated": "Y" if found else "N",
                "normalized_populated": "Y" if found else "N",
                "core_metric_completion_pct": round(sum(1 for f in CORE_FIELDS if found and found.get(f) not in ("", None)) / len(CORE_FIELDS), 3) if found else 0,
                "gap_note": "" if found else "No full-quarter 2026 Q2 earnings source included in Phase 1B; stop line remains latest verified full quarter.",
            })
    fedex_periods = [
        ("FY2024 Q1", "2023-08-31"), ("FY2024 Q2", "2023-11-30"),
        ("FY2024 Q3", "2024-02-29"), ("FY2024 Q4", "2024-05-31"), ("FY2025 Q1", "2024-08-31"),
        ("FY2025 Q2", "2024-11-30"), ("FY2025 Q3", "2025-02-28"), ("FY2025 Q4", "2025-05-31"),
        ("FY2026 Q1", "2025-08-31"), ("FY2026 Q2", "2025-11-30"), ("FY2026 Q3", "2026-02-28"),
        ("FY2026 Q4", "2026-05-31"),
    ]
    for label, end in fedex_periods:
        found = next((r for r in rows if r["ticker"] == "FDXF" and r["period_end_date"] == end), None)
        expected.append({
            "carrier": "FedEx Freight", "ticker": "FDXF", "period_label": label, "period_end_date": end,
            "expected": "Y", "source_found": "Y" if found else "N", "downloaded": found["source_status"] if found else "N",
            "raw_populated": "Y" if found else "N",
            "normalized_populated": "Y" if found else "N",
            "core_metric_completion_pct": round(sum(1 for f in CORE_FIELDS if found and found.get(f) not in ("", None)) / len(CORE_FIELDS), 3) if found else 0,
            "gap_note": "" if found else "Quarter-level FedEx Freight table still unresolved for pre-FY2026 period.",
        })
    return expected


def normalized_rows():
    comparable = [
        "observation_id", "carrier", "ticker", "period_end_date", "calendar_year", "calendar_quarter",
        "fiscal_year_reported", "fiscal_quarter_reported", "comparison_alignment", "revenue_m",
        "operating_income_m", "adjusted_operating_income_m", "operating_margin_pct",
        "adjusted_operating_margin_pct", "operating_ratio_pct", "adjusted_operating_ratio_pct",
        "avg_daily_shipments", "shipments_yoy_pct", "tonnage_per_day", "tonnage_yoy_pct",
        "weight_per_shipment_lbs", "revenue_per_shipment", "revenue_per_shipment_ex_fuel",
        "revenue_per_cwt", "revenue_per_cwt_ex_fuel", "revenue_per_cwt_ex_fuel_yoy_pct",
        "base_revenue_per_cwt_yoy_pct", "comparability", "comparability_note", "source_key",
    ]
    return [{h: r.get(h, "") for h in comparable} for r in rows], comparable


def service_rows():
    ranking_rows = [
        ("Old Dominion Freight Line", "ODFL", 2021, 1, "", "MASTIO_2021_ODFL", "CARRIER_IR_RELEASE_QUOTING_MASTIO", "Official carrier release confirms ODFL #1; denominator and peer ranks not publicly retrieved."),
        ("Old Dominion Freight Line", "ODFL", 2022, 1, "", "MASTIO_2022_ODFL", "CARRIER_IR_RELEASE_QUOTING_MASTIO", "Official carrier release confirms ODFL #1; denominator and peer ranks not publicly retrieved."),
        ("Old Dominion Freight Line", "ODFL", 2023, 1, 7, "MASTIO_2023_ODFL", "CARRIER_IR_RELEASE_QUOTING_MASTIO", "ODFL #1 supported by official carrier release; denominator from Phase 1C request."),
        ("Saia", "SAIA", 2023, 3, 7, "MASTIO_2023_TRADE", "TRADE_PUBLICATION_PLUS_USER_SPECIFIED_DENOMINATOR", "Rank from Phase 1C request; trade source confirms Saia top-five national rank."),
        ("XPO North American LTL", "XPO", 2023, 4, 7, "MASTIO_2023_TRADE", "TRADE_PUBLICATION_PLUS_USER_SPECIFIED_DENOMINATOR", "Rank from Phase 1C request; trade source confirms XPO top-five national rank."),
        ("FedEx Freight", "FDXF", 2023, 6, 7, "MASTIO_2023_TRADE", "USER_PROVIDED_PENDING_OFFICIAL", "Rank from Phase 1C request; official Mastio 19th Edition full press release not publicly retrieved."),
        ("Old Dominion Freight Line", "ODFL", 2024, 1, "", "MASTIO_2024_TRADE", "TRADE_PUBLICATION", "FreightWaves reports ODFL top national carrier; denominator not publicly retrieved."),
        ("XPO North American LTL", "XPO", 2024, 4, "", "MASTIO_2024_TRADE", "TRADE_PUBLICATION", "FreightWaves reports XPO held national rank #4; denominator not publicly retrieved."),
        ("Saia", "SAIA", 2024, 5, "", "MASTIO_2024_TRADE", "TRADE_PUBLICATION", "FreightWaves reports Saia national rank #5; denominator not publicly retrieved."),
        ("FedEx Freight", "FDXF", 2024, "", "", "MASTIO_2024_TRADE", "NOT_AVAILABLE", "FedEx exact national-carrier rank not publicly retrieved; do not infer."),
        ("Old Dominion Freight Line", "ODFL", 2025, 1, 8, "MASTIO_2025_ODFL", "CARRIER_IR_RELEASE_QUOTING_MASTIO", "ODFL #1 official carrier release; denominator from public 21st Edition summary."),
        ("XPO North American LTL", "XPO", 2025, 4, 8, "MASTIO_2025_SECONDARY", "SECONDARY_PUBLIC_SUMMARY", "Public summary of 21st Edition national-carrier ranking."),
        ("Saia", "SAIA", 2025, 6, 8, "MASTIO_2025_SECONDARY", "SECONDARY_PUBLIC_SUMMARY", "Public summary of 21st Edition national-carrier ranking."),
        ("FedEx Freight", "FDXF", 2025, 7, 8, "MASTIO_2025_SECONDARY", "SECONDARY_PUBLIC_SUMMARY", "Public summary of 21st Edition national-carrier ranking."),
    ]
    out = []
    for carrier, ticker, year, rank, denominator, key, quality, note in ranking_rows:
        src = SOURCES[key]
        percentile = ""
        if rank != "" and denominator not in ("", 1):
            percentile = round((denominator - rank) / (denominator - 1), 4)
        out.append({
            "carrier": carrier,
            "ticker": ticker,
            "year": year,
            "service_measure": "Mastio National Carrier ranking",
            "rank": rank,
            "rank_denominator": denominator,
            "normalized_percentile": percentile,
            "overall_ranking": "",
            "source_quality": quality,
            "source_publication_date": src["publication_date"],
            "source_url": src["url"],
            "local_source_file": str(SRC_DIR / src["file"]) if src.get("file") else "",
            "notes": note,
        })
    return out


def dictionary_rows():
    return [
        ("observation_id", "Stable row id composed from ticker, fiscal period, and period end date.", "text", "", "n/a", "Generated", "Used to join provenance rows."),
        ("carrier", "Carrier or reporting segment name.", "text", "", "n/a", "Company source", ""),
        ("period_end_date", "Reported quarter end date.", "date", "YYYY-MM-DD", "n/a", "Company source", "FedEx uses fiscal quarter ends."),
        ("comparison_period", "Prior period used by source for year-over-year comparison.", "text", "", "n/a", "Company source", "Can differ from calendar peers for FedEx."),
        ("comparison_alignment", "EXACT if current/prior workday quarter aligns; APPROXIMATE when workdays/fiscal alignment differ; NOT_ALIGNED for prior-column extraction without current growth.", "text", "", "n/a", "Analyst classification", ""),
        ("revenue_m", "Quarterly operating revenue or segment revenue.", "number", "USD millions", "Higher", "Company source", "Saia/ODFL source tables are in thousands and converted to millions."),
        ("operating_income_m", "Quarterly GAAP operating income.", "number", "USD millions", "Higher", "Company source", "Converted to millions when reported in thousands."),
        ("adjusted_operating_income_m", "Quarterly adjusted operating income when reported.", "number", "USD millions", "Higher", "Company non-GAAP source", "Blank when not explicitly reported."),
        ("operating_margin_pct", "Operating income divided by revenue, or source-reported operating margin.", "number", "percent", "Higher", "Company source or algebraic complement", "For ODFL/Saia derived from operating ratio where margin not separately reported."),
        ("operating_ratio_pct", "Operating expenses as a percent of revenue, or 100 minus operating margin.", "number", "percent", "Lower", "Company source or algebraic complement", ""),
        ("adjusted_operating_ratio_pct", "Adjusted operating ratio where reported.", "number", "percent", "Lower", "Company non-GAAP source", "XPO reports adjusted OR; FedEx reports adjusted margin."),
        ("avg_daily_shipments", "Average daily LTL shipments.", "number", "shipments/day", "Context dependent", "Company source", "Saia and FedEx thousand-scale values multiplied by 1,000."),
        ("shipments_yoy_pct", "Year-over-year change in average daily shipments.", "number", "percent", "Context dependent", "Company source", ""),
        ("tonnage_per_day", "Average daily tonnage.", "number", "tons/day", "Context dependent", "Company source or unit conversion", "XPO pounds/day in thousands converted to tons/day."),
        ("revenue_per_shipment_ex_fuel", "Revenue per shipment excluding fuel surcharge.", "number", "USD/shipment", "Higher", "Company source", "Blank when not reported."),
        ("revenue_per_cwt_ex_fuel", "Revenue per hundredweight excluding fuel surcharge.", "number", "USD/CWT", "Higher", "Company source", "Blank when not reported."),
        ("base_revenue_per_cwt_yoy_pct", "Base revenue per hundredweight year-over-year change.", "number/text", "percent or directional text", "Higher", "Company source", "FedEx FY2026 Q4 only gives directional 'slight decline'."),
        ("source_publication_date", "Publication date of the cited release or exhibit.", "date", "YYYY-MM-DD", "n/a", "Source metadata", "No placeholder values allowed."),
        ("exact_reported_metric_label", "Source table or narrative label used for the datapoint.", "text", "", "n/a", "Manual extraction", "Populated in Provenance for each datapoint."),
    ]


def validation_flags():
    flags = []
    for r in rows:
        if r["comparison_alignment"] != "EXACT":
            flags.append([r["observation_id"], r["carrier"], r["period_end_date"], "ALIGNMENT", r["comparison_alignment"], r["comparability_note"]])
        if r["source_status"] != "DOWNLOADED":
            flags.append([r["observation_id"], r["carrier"], r["period_end_date"], "SOURCE_DOWNLOAD", r["source_status"], "URL retained; local download unavailable or timed out."])
        if r.get("base_revenue_per_cwt_yoy_pct") == "slight decline":
            flags.append([r["observation_id"], r["carrier"], r["period_end_date"], "DIRECTIONAL_METRIC", "No numeric value", "Source reports base revenue/CWT directionally only."])
    return flags


def unresolved_gaps():
    return [
        ["FedEx Freight", "FY2024-FY2025", "Backfilled from FedEx historical statistical book; local official FedEx investor download was blocked by Cloudflare, so source URL points to a public mirror of FedEx's statistical book content.", "Monitor", "Replace mirror URL with direct FedEx investor PDF/XLSX if accessible."],
        ["FedEx Freight", "FY2026 Q4", "Base revenue per CWT reported only as a slight decline, not a numeric percentage.", "Open", "Numeric base pricing fields remain blank except directional note."],
        ["Mastio service", "2021-2022", "Only ODFL #1 official carrier releases were found; full national-carrier rank tables were not retrieved.", "Open", "Do not infer unavailable peer ranks."],
        ["Mastio service", "2024", "Trade source gives ODFL/XPO/Saia national ranks but FedEx exact rank and denominator were not retrieved.", "Open", "Do not infer unavailable FedEx rank."],
        ["Mastio service", "2025 non-ODFL ranks", "Official ODFL/Mastio release confirms ODFL #1; non-ODFL national ranks come from public summary.", "Monitor", "Replace with official Mastio 21st Edition full release if obtained."],
        ["XPO", "2024-2025", "IR source pages found, but local HTML downloads timed out before completion.", "Open", "URLs retained in source_url; Q1 2026 SEC exhibit downloaded locally."],
    ]


def manual_verification(prov):
    samples = [
        ("ODFL", "2024-03-31", "avg_daily_shipments", 46931),
        ("ODFL", "2024-06-30", "operating_ratio_pct", 71.9),
        ("ODFL", "2025-09-30", "revenue_per_cwt_ex_fuel", 28.78),
        ("ODFL", "2026-03-31", "weight_per_shipment_lbs", 1491),
        ("SAIA", "2024-06-30", "revenue_m", 823.244),
        ("SAIA", "2024-06-30", "revenue_per_shipment_ex_fuel", 290.72),
        ("SAIA", "2025-06-30", "avg_daily_shipments", 35330),
        ("SAIA", "2026-03-31", "revenue_per_cwt_ex_fuel", 21.52),
        ("XPO", "2024-03-31", "revenue_m", 1221),
        ("XPO", "2024-06-30", "tonnage_per_day", 36329.0),
        ("XPO", "2024-12-31", "revenue_per_cwt_ex_fuel", 24.84),
        ("XPO", "2025-03-31", "load_factor", 22434),
        ("XPO", "2025-09-30", "adjusted_operating_ratio_pct", 82.7),
        ("XPO", "2025-12-31", "avg_daily_shipments", 48348),
        ("XPO", "2026-03-31", "revenue_per_cwt_ex_fuel", 25.71),
        ("FDXF", "2025-11-30", "revenue_m", 2139),
        ("FDXF", "2025-11-30", "avg_daily_shipments", 87438),
        ("FDXF", "2026-05-31", "revenue_m", 2400),
        ("FDXF", "2026-05-31", "adjusted_operating_margin_pct", 15.1),
        ("FDXF", "2026-05-31", "revenue_per_cwt", 43.79),
    ]
    out = []
    for i, (ticker, end, field, expected) in enumerate(samples, start=1):
        r = next(x for x in rows if x["ticker"] == ticker and x["period_end_date"] == end)
        out.append({
            "check_id": f"MV{i:02d}",
            "carrier": r["carrier"],
            "period_end_date": end,
            "field_name": field,
            "workbook_value": r.get(field, ""),
            "expected_value_from_source": expected,
            "status": "PASS" if r.get(field, "") == expected else "REVIEW",
            "source_key": r["source_key"],
            "notes": "Verified against parsed table/narrative capture in Phase 1B source inventory.",
        })
    return out


def company_service_kpis():
    return [
        {"carrier": "Old Dominion Freight Line", "ticker": "ODFL", "period_end_date": "2026-06-30", "metric": "on_time_service_pct", "value": 99.0, "qualifier": "above", "source_key": "ODFL_2026_Q2", "notes": "Company-reported quarterly service KPI; separate from Mastio."},
        {"carrier": "Old Dominion Freight Line", "ticker": "ODFL", "period_end_date": "2026-06-30", "metric": "cargo_claims_ratio_pct", "value": 0.1, "qualifier": "", "source_key": "ODFL_2026_Q2", "notes": "Company-reported quarterly service KPI; separate from Mastio."},
        {"carrier": "XPO North American LTL", "ticker": "XPO", "period_end_date": "2026-06-30", "metric": "damage_claims_ratio_pct", "value": 0.2, "qualifier": "less than", "source_key": "XPO_2026_Q2", "notes": "Company-reported quarterly service KPI; separate from Mastio."},
        {"carrier": "Saia", "ticker": "SAIA", "period_end_date": "2026-06-30", "metric": "claims_ratio_pct", "value": 0.3, "qualifier": "", "source_key": "SAIA_2026_Q2", "notes": "Company-reported quarterly service KPI; separate from Mastio."},
    ]


def fedex_priority_economy_rows():
    specs = [
        ("FY2024", "Q1", "2023-08-31", 66144, 28491, 989, 876, 353.01, 407.99, 35.71, 46.59),
        ("FY2024", "Q2", "2023-11-30", 68486, 30515, 975, 880, 365.55, 415.82, 37.48, 47.26),
        ("FY2024", "Q3", "2024-02-29", 61483, 27765, 974, 885, 363.21, 414.79, 37.31, 46.89),
        ("FY2024", "Q4", "2024-05-31", 63556, 29524, 970, 871, 364.11, 406.73, 37.52, 46.70),
        ("FY2025", "Q1", "2024-08-31", 62893, 29115, 956, 868, 363.97, 408.60, 38.06, 47.09),
        ("FY2025", "Q2", "2024-11-30", 62513, 28485, 935, 865, 352.84, 400.00, 37.73, 46.26),
        ("FY2025", "Q3", "2025-02-28", 58186, 26886, 935, 877, 360.68, 408.56, 38.57, 46.59),
        ("FY2025", "Q4", "2025-05-31", 63397, 28732, 937, 882, 357.86, 404.98, 38.18, 45.92),
        ("FY2026", "Q1", "2025-08-31", 62029, 27983, 947, 876, 362.63, 401.20, 38.31, 45.78),
        ("FY2026", "Q2", "2025-11-30", 60138, 27300, 933, 904, 360.64, 409.75, 38.67, 45.35),
        ("FY2026", "Q3", "2026-02-28", 55588, 24645, 943, 887, 364.37, 416.05, 38.64, 46.92),
    ]
    out = []
    for fy, fq, end, p_ship, e_ship, p_weight, e_weight, p_rps, e_rps, p_cwt, e_cwt in specs:
        out.append({"fiscal_year": fy, "fiscal_quarter": fq, "period_end_date": end, "service_level": "Priority", "avg_daily_shipments": p_ship, "weight_per_shipment_lbs": p_weight, "revenue_per_shipment": p_rps, "revenue_per_cwt": p_cwt, "source_key": "FDX_STATBOOK_FY2026_Q3", "notes": "FedEx Freight segment quarterly operating statistics supplement."})
        out.append({"fiscal_year": fy, "fiscal_quarter": fq, "period_end_date": end, "service_level": "Economy", "avg_daily_shipments": e_ship, "weight_per_shipment_lbs": e_weight, "revenue_per_shipment": e_rps, "revenue_per_cwt": e_cwt, "source_key": "FDX_STATBOOK_FY2026_Q3", "notes": "FedEx Freight segment quarterly operating statistics supplement."})
    return out


def write_sheet(ws, data, headers=None):
    if not data and not headers:
        return
    if headers is None:
        headers = list(data[0].keys())
    ws.append(headers)
    for item in data:
        if isinstance(item, dict):
            ws.append([item.get(h, "") for h in headers])
        else:
            ws.append(item)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 48)
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_csv(name, data, headers):
    path = OUT_DIR / name
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(data)


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    prov = build_provenance()
    comp = completeness_matrix()
    norm, norm_headers = normalized_rows()
    service = service_rows()
    flags = validation_flags()
    gaps = unresolved_gaps()
    mv = manual_verification(prov)

    wb = Workbook()
    wb.remove(wb.active)

    readme = wb.create_sheet("README")
    write_sheet(readme, [
        ["Workbook", "LTL peer panel Phase 1C final repair - FedEx backfill, Q2 2026 peers, Mastio history"],
        ["Scope lock", "Full peer panel remains 2024-2026. FedEx FY2024 historical rows are included only to repair FedEx Freight statistical-book backfill."],
        ["Status", "Final critical data repair workbook; gaps are explicit in Quarter_Completeness and Data_Gaps."],
        ["Thesis conclusions", "No causation claim or stock recommendation included by request."],
        ["Created", date.today().isoformat()],
    ], ["item", "detail"])

    write_sheet(wb.create_sheet("Quarter_Completeness"), comp)

    dash_rows = []
    for ticker in sorted({r["ticker"] for r in rows}):
        carrier_rows = [r for r in rows if r["ticker"] == ticker]
        for field in CORE_FIELDS:
            dash_rows.append({
                "ticker": ticker,
                "field_name": field,
                "populated_rows": sum(1 for r in carrier_rows if r.get(field) not in ("", None)),
                "total_extracted_rows": len(carrier_rows),
                "completion_pct": round(sum(1 for r in carrier_rows if r.get(field) not in ("", None)) / len(carrier_rows), 3),
            })
    write_sheet(wb.create_sheet("Completeness_Dashboard"), dash_rows)

    write_sheet(wb.create_sheet("Quarterly_Raw"), rows, RAW_HEADERS)
    write_sheet(wb.create_sheet("Quarterly_Normalized"), norm, norm_headers)
    write_sheet(wb.create_sheet("Service_Annual"), service)
    write_sheet(wb.create_sheet("Company_Service_KPI"), company_service_kpis())
    write_sheet(wb.create_sheet("FedEx_Priority_Economy"), fedex_priority_economy_rows())
    write_sheet(wb.create_sheet("Network_Annual"), [
        {"carrier": "Old Dominion Freight Line", "ticker": "ODFL", "year": 2026, "metric": "average_active_full_time_employees", "value": 20264, "source": "ODFL 2026 Q1 release", "notes": "Quarterly source metric retained here pending annual network build."},
    ])
    write_sheet(wb.create_sheet("Provenance"), prov)
    write_sheet(wb.create_sheet("Data_Dictionary"), dictionary_rows(), ["field_name", "definition", "type", "unit", "favorable_direction", "source_or_derivation", "notes"])
    write_sheet(wb.create_sheet("Validation_Flags"), flags, ["observation_id", "carrier", "period_end_date", "flag_type", "flag_value", "notes"])
    write_sheet(wb.create_sheet("Manual_Verification"), mv)
    write_sheet(wb.create_sheet("Data_Gaps"), gaps, ["carrier", "period", "gap", "status", "next_step"])
    write_sheet(wb.create_sheet("Analysis"), [
        ["A. Persistent FedEx Mastio disadvantage?", "In the verifiable full-rank years entered, yes: FedEx Freight ranks 6/7 in 2023 and 7/8 in 2025, while ODFL is consistently #1. 2024 does not have an exact FedEx rank in the public sources found, so persistence is supported but not complete for every year."],
        ["B. Shipment growth comparison", "FedEx Freight shipment growth is mostly negative across FY2024-FY2026. In the latest comparable quarter, FDXF FY2026 Q4 shipments were -5.9%; ODFL Q2 2026 was -5.7%, XPO Q2 2026 was +2.8%, and Saia Q2 2026 was +4.4%."],
        ["C. Service/economics coincidence", "Weaker service appears to coincide more consistently with weaker volume than weaker pricing. FDXF pricing per shipment/CWT improved in FY2026 Q4 even while shipments declined. ODFL pairs top service with strong yield but negative shipments; XPO and Saia show positive Q2 2026 shipments with mixed service/pricing indicators."],
        ["D. Contradictory evidence", "FedEx Freight showed strong FY2026 Q4 revenue per shipment and CWT despite low Mastio rank; ODFL maintained #1 service while shipments declined; Saia gained Q2 2026 volume even with weaker ex-fuel CWT; XPO's middle-ranked Mastio position did not prevent strong Q2 2026 volume and pricing."],
        ["Causation / recommendation", "No causation claim and no stock recommendation are made."],
    ], ["question", "answer"])

    wb.save(XLSX)

    write_csv("quarterly_raw.csv", rows, RAW_HEADERS)
    write_csv("quarterly_normalized.csv", norm, norm_headers)
    write_csv("service_annual.csv", service, list(service[0].keys()))
    write_csv("provenance.csv", prov, list(prov[0].keys()))

    # Verification pass: confirm file opens and expected sheets exist.
    test = load_workbook(XLSX, read_only=True, data_only=False)
    expected_sheets = {
        "README", "Quarter_Completeness", "Completeness_Dashboard", "Quarterly_Raw",
        "Quarterly_Normalized", "Service_Annual", "Company_Service_KPI", "FedEx_Priority_Economy", "Network_Annual", "Provenance",
        "Data_Dictionary", "Validation_Flags", "Manual_Verification", "Data_Gaps", "Analysis",
    }
    missing = expected_sheets.difference(test.sheetnames)
    if missing:
        raise RuntimeError(f"Missing sheets: {sorted(missing)}")
    test.close()

    print(f"Wrote {XLSX}")
    print(f"Raw rows: {len(rows)}; provenance rows: {len(prov)}; manual checks: {len(mv)}")


if __name__ == "__main__":
    main()
