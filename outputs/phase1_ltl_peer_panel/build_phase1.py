from pathlib import Path
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

out = Path("outputs/phase1_ltl_peer_panel")
out.mkdir(parents=True, exist_ok=True)

q_headers = ["observation_id","carrier","ticker","period_start_date","period_end_date","calendar_year","calendar_quarter","fiscal_year_reported","fiscal_quarter_reported","operating_days","reporting_basis","revenue_m","operating_income_m","adjusted_operating_income_m","operating_margin_pct","adjusted_operating_margin_pct","operating_ratio_pct","adjusted_operating_ratio_pct","avg_daily_shipments","shipments_yoy_pct","tonnage_per_day","tonnage_yoy_pct","weight_per_shipment_lbs","weight_per_shipment_yoy_pct","revenue_per_shipment","revenue_per_shipment_yoy_pct","revenue_per_shipment_ex_fuel","revenue_per_shipment_ex_fuel_yoy_pct","revenue_per_cwt","revenue_per_cwt_yoy_pct","revenue_per_cwt_ex_fuel","revenue_per_cwt_ex_fuel_yoy_pct","yield_reported","yield_ex_fuel_reported","fuel_surcharge_revenue_m","fuel_surcharge_pct_revenue","comparability","comparability_note","source_url","local_source_file","source_status"]
rows = [
["FDXF_FY2026_Q4","FedEx Freight","FDXF","2026-03-01","2026-05-31",2026,"Q2",2026,"Q4",None,"FedEx Freight segment as reported by FedEx Corp; not carve-out basis",2400,158,363,0.066,0.151,None,None,86700,-0.059,None,None,948,0.03,415.22,0.115,None,None,43.79,0.082,None,None,None,None,None,None,"PARTIAL","FedEx fiscal quarter ending May 31; includes fuel for revenue/shipment and revenue/cwt; peer comparability requires basis review.","https://ir.fedexfreight.com/news-events/press-releases/detail/185/fedex-freight-reports-fourth-quarter-and-full-fiscal-year-2026-financial-results","FDXF_2026_Q4_earnings_release.html","Downloaded"],
["ODFL_CY2026_Q1","Old Dominion Freight Line","ODFL","2026-01-01","2026-03-31",2026,"Q1",2026,"Q1",63,"Reported company LTL statistics / GAAP operating ratio",None,None,None,None,None,0.762,None,41037,-0.079,30584,-0.077,1491,0.003,514.56,0.059,434.16,0.047,34.52,0.057,29.13,0.044,None,None,None,None,"HIGH","Calendar quarter; ODFL provides absolute operating statistics and ex-fuel pricing metrics.","https://ir.odfl.com/news-events/press-releases/detail/342/old-dominion-freight-line-reports-first-quarter-2026","ODFL_2026_Q1_earnings_release.html","Downloaded"],
["XPO_CY2026_Q1","XPO North American LTL","XPO","2026-01-01","2026-03-31",2026,"Q1",2026,"Q1",None,"North American LTL segment",1230,189,198,None,None,None,0.839,None,0.03,None,0.001,None,None,None,None,None,None,None,None,None,0.04,None,0.04,None,None,"PARTIAL","XPO release reports segment revenue/income plus YoY shipment, tonnage, and yield changes; absolute shipment/tonnage not in captured release snippet.","https://investors.xpo.com/news-releases/news-release-details/xpo-reports-first-quarter-2026-results","XPO_2026_Q1_earnings_release.html","Identified, download timed out"],
["SAIA_CY2026_Q1","Saia","SAIA","2026-01-01","2026-03-31",2026,"Q1",2026,"Q1",63,"GAAP company LTL operating statistics",806.2,66.8,None,None,None,0.917,None,34790,0.01,24020,-0.021,1380,None,None,None,None,-0.012,25.93,0.038,21.52,0.019,None,None,None,None,"HIGH","Calendar quarter; LTL stats exclude transportation/logistics services where pricing is not generally determined by weight.","https://www.sec.gov/Archives/edgar/data/1177702/000119312526194062/saia-ex99_1.htm","SAIA_2026_Q1_SEC_8K_ex99_1.htm","Downloaded"],
["SAIA_CY2025_Q2","Saia","SAIA","2025-04-01","2025-06-30",2025,"Q2",2025,"Q2",64,"GAAP company LTL operating statistics",817.1,99.4,None,None,None,0.878,None,35330,-0.028,24630,0.011,1394,0.04,351.36,0.018,298.71,0.027,25.20,-0.021,21.42,-0.012,None,None,None,None,"HIGH","Downloaded SEC exhibit includes financial and operating table for Q2 2025 and prior-year comparison.","https://www.sec.gov/Archives/edgar/data/1177702/000095017025098577/saia-ex99_1.htm","SAIA_2025_Q2_SEC_8K_ex99_1.htm","Downloaded"],
["XPO_CY2025_Q2","XPO North American LTL","XPO","2025-04-01","2025-06-30",2025,"Q2",2025,"Q2",None,"North American LTL segment",1240,199,211,None,None,None,0.829,None,-0.051,None,-0.067,None,None,None,0.056,None,None,None,None,None,None,0.042,0.061,None,None,"PARTIAL","Release reports segment financials and YoY operating/yield changes; absolute shipment/tonnage not captured in first pass.","https://investors.xpo.com/news-releases/news-release-details/xpo-reports-second-quarter-2025-results","XPO_2025_Q2_earnings_release.html","Identified, download timed out"],
]

service_headers = ["carrier","year","mastio_rank","mastio_rank_denominator","mastio_percentile","mastio_overall_rank","mastio_national_carrier_rank","mastio_overall_weighted_quality_score","on_time_service_pct","cargo_claims_ratio_pct","damage_claims_ratio_pct","service_metric_definition","source_url","source_status","notes"]
service_rows = [[c,y,None,None,None,None,None,None,None,None,None,None,None,"Gap","Mastio/public service source pending"] for y in (2024,2025) for c in ("FedEx Freight","Old Dominion Freight Line","XPO North American LTL","Saia")]

prov_headers = ["observation_id","carrier","metric_name","metric_value","unit","period","source_type","source_title","source_url","source_publication_date","source_page_or_section","exact_reported_metric_label","reported_or_derived","derivation_formula","notes"]
prov_rows = []
metrics = ["revenue_m","operating_income_m","adjusted_operating_income_m","operating_ratio_pct","avg_daily_shipments","shipments_yoy_pct","tonnage_per_day","tonnage_yoy_pct","weight_per_shipment_lbs","revenue_per_shipment","revenue_per_cwt","revenue_per_cwt_ex_fuel","yield_ex_fuel_reported"]
for r in rows:
    d = dict(zip(q_headers, r))
    for m in metrics:
        if d.get(m) not in (None, ""):
            prov_rows.append([d["observation_id"], d["carrier"], m, d[m], "decimal" if ("pct" in m or "ratio" in m or "yield" in m) else "reported units", f'{d["calendar_year"]} {d["calendar_quarter"]}', "Primary company / SEC source", "Phase 1 captured release", d["source_url"], "See source", "Press release / operating statistics", m, "REPORTED", "", "First pass; page/section to be tightened during full extraction."])

dict_rows = [[h, "Requested Phase 1 field. Percent fields use decimal values; blanks mean not source-backed in this pass."] for h in q_headers]
gaps = [["Area","Gap","Priority","Next step"],["Downloads","XPO 2026 Q1 and 2025 Q2 identified but local download timed out","High","Retry company PDF links or SEC exhibits"],["2024 coverage","No 2024 rows populated yet in first pass","High","Download 2024 quarterly releases/stat books before backward expansion"],["Mastio","2024-2025 official Mastio ranks/denominators not populated","High","Locate official Mastio releases/methodology"],["FedEx","Need FY2024-FY2026 stat books for full quarter detail","High","Use FedEx IR stat book PDFs/XLSX"],["Service KPIs","On-time and claims definitions mostly blank","Medium","Populate only when publicly disclosed"]]

wb = Workbook()
wb.remove(wb.active)
tabs = {
    "README":[["Phase 1 preliminary workbook"],["Scope: identify/download primary sources, create data dictionary, populate 2024-2026 first."],["Do not proceed to 2021-2023 until this workbook and gaps are reviewed."]],
    "Quarterly_Raw":[q_headers]+rows,
    "Quarterly_Normalized":[q_headers]+[r[:] for r in rows],
    "Service_Annual":[service_headers]+service_rows,
    "Network_Annual":[["carrier","year","service_center_count","door_count","tractor_count","trailer_count","employee_count","driver_count","capex_m","source_url","source_status","notes"]],
    "Provenance":[prov_headers]+prov_rows,
    "Data_Dictionary":[["field_name","definition"]]+dict_rows,
    "Validation_Flags":gaps,
    "Analysis":[["Preliminary Readout"],["Current data are too sparse for charts or conclusions."],["Populated rows"],["FedEx Freight",1],["ODFL",1],["XPO",2],["Saia",2]],
    "Charts":[["Charts intentionally deferred"],["Need fuller 2024-2026 coverage before trend plots."]],
}
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="B7C9D6")
for name, data in tabs.items():
    ws = wb.create_sheet(name)
    for row in data:
        ws.append(row)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col in ws.columns:
        width = min(max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2, 45)
        ws.column_dimensions[get_column_letter(col[0].column)].width = width

norm = wb["Quarterly_Normalized"]
for row in range(2, len(rows)+2):
    norm[f"O{row}"] = f'=IF(Q{row}<>"",1-Q{row},IF(L{row}<>"",M{row}/L{row},""))'
for ws in (wb["Quarterly_Raw"], wb["Quarterly_Normalized"]):
    for col in ("O","P","Q","R","T","V","X","Z","AB","AD","AF","AH","AJ"):
        for row in range(2, len(rows)+2):
            ws[f"{col}{row}"].number_format = "0.0%"

for name, headers, data in [
    ("quarterly_raw.csv", q_headers, rows),
    ("quarterly_normalized.csv", q_headers, rows),
    ("service_annual.csv", service_headers, service_rows),
    ("provenance.csv", prov_headers, prov_rows),
]:
    with (out/name).open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(data)

wb.save(out/"LTL_peer_panel_PHASE1_2024_2026_preliminary.xlsx")
print(out/"LTL_peer_panel_PHASE1_2024_2026_preliminary.xlsx")
