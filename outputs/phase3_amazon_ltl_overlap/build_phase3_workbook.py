from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule


OUT_DIR = Path(__file__).resolve().parent
WORKBOOK = OUT_DIR / "Amazon_LTL_FDXF_Overlap.xlsx"


SOURCE = {
    "AMZN26_PR": "Amazon Press Center, Jun. 10, 2026, Amazon Supply Chain Services launches LTL freight offering for all businesses",
    "AMZN26_LTL": "Amazon Freight newsroom, Jun. 10, 2026, Amazon Freight now offers LTL for shippers of any size",
    "AMZN26_PAGE": "Amazon Freight LTL service page, 2026 coverage map and service proposition",
    "AMZN25_IN": "Amazon Freight newsroom, Apr. 4, 2025, Amazon Freight launches inbound LTL for Amazon facilities",
    "AMZN25_BOOK": "Amazon Freight newsroom, Sep. 10, 2025, inbound LTL booking workflow",
    "AMZN24_FAQ": "Amazon Freight newsroom, Jan. 19, 2024, top Amazon Freight questions",
    "AMZN_HOME": "Amazon Freight home page, carrier model, quote tools, asset-backed broker description",
    "FDXF_FORM10": "FedEx Freight Form 10 Exhibit 99.1, network and terminal disclosure",
    "FDXF_INV": "FedEx Freight Investor Day presentation, Jun. 2026, strategy and go-to-market slides",
    "FDXF_PAGE": "FedEx Freight public page, national LTL positioning and digital tools",
    "XPO_PAGE": "XPO geographic coverage and service center pages",
    "ODFL_PAGE": "Old Dominion about/service-center pages and Q1 2026 release",
    "SAIA_10K": "Saia 2025 Form 10-K, network and equipment disclosure",
}


URL = {
    "AMZN26_PR": "https://press.aboutamazon.com/2026/6/amazon-supply-chain-services-launches-less-than-truckload-freight-offering-for-all-businesses",
    "AMZN26_LTL": "https://freight.amazon.com/newsroom/news-updates/2026-ltl-launch",
    "AMZN26_PAGE": "https://freight.amazon.com/services/less-than-truckload",
    "AMZN25_IN": "https://freight.amazon.com/newsroom/2025-ltl-inbound-launch",
    "AMZN25_BOOK": "https://freight.amazon.com/newsroom/news-updates/news-updates-25-inbound-ltl-booking",
    "AMZN24_FAQ": "https://freight.amazon.com/newsroom/2024-af-top-questions",
    "AMZN_HOME": "https://freight.amazon.com/",
    "FDXF_FORM10": "https://ir.fedexfreight.com/financial-information/sec-filings/content/0001104659-26-060223/tm2520565d14_ex99-1.htm",
    "FDXF_INV": "https://d1io3yog0oux5.cloudfront.net/_5705b101ff53031529c842d008f96e5d/fedexfreight/db/3570/34491/presentation/FedEx+Freight+Investor+Day+Presentation.pdf",
    "FDXF_PAGE": "https://www.fedexfreight.com/en-us",
    "XPO_PAGE": "https://www.xpo.com/offerings/geographic-coverage/",
    "ODFL_PAGE": "https://www.olddominionfreightlinesandservices.com/us/en/tools/service-center-locator.html",
    "SAIA_10K": "https://www.sec.gov/Archives/edgar/data/1177702/000119312526067030/saia-20251231.htm",
}


def score_label(score):
    if score >= 0.8:
        return "HIGH"
    if score >= 0.5:
        return "MEDIUM"
    if score > 0:
        return "LOW"
    return "NONE"


def src_urls(source_ids):
    ids = [sid.strip() for sid in source_ids.split(";")]
    return "\n".join(URL[sid] for sid in ids if sid in URL)


REGION = {
    "Dallas/Fort Worth": "Texas",
    "Houston": "Texas",
    "San Antonio/Austin": "Texas",
    "Los Angeles/Inland Empire": "West",
    "San Francisco/Oakland": "West",
    "Sacramento": "West",
    "Seattle": "West",
    "Portland": "West",
    "Salt Lake City": "Mountain West",
    "Las Vegas": "West",
    "Chicago": "Midwest",
    "Milwaukee": "Midwest",
    "Minneapolis/St. Paul": "Midwest",
    "Detroit": "Great Lakes",
    "Cleveland/North Jackson": "Great Lakes",
    "Pittsburgh": "Great Lakes",
    "Columbus": "Great Lakes",
    "Cincinnati/Dayton": "Great Lakes",
    "Indianapolis": "Midwest",
    "Louisville": "South",
    "Nashville": "South",
    "Memphis": "South",
    "Kansas City": "Central",
    "St. Louis": "Central",
    "Atlanta": "Southeast",
    "Charlotte": "Southeast",
    "Raleigh": "Southeast",
    "Jacksonville": "Florida",
    "Orlando/Tampa": "Florida",
    "Miami": "Florida",
    "New York/Northern New Jersey": "Northeast",
    "Philadelphia": "Northeast",
    "Baltimore/DC/Hagerstown": "Mid-Atlantic",
    "Boston": "Northeast",
    "Denver": "Mountain West",
    "Phoenix": "Southwest",
    "Des Moines": "Midwest",
}


timeline = [
    ["2019-2024", "Amazon-linked freight and early LTL flows", "Amazon selling partners, vendors and businesses using Amazon Freight; public 2024 FAQ emphasized FTL as the core service", "Primarily Amazon-related inbound/vendor flows; FTL available to external shippers", "Lower 48 in areas where Amazon operates; not disclosed as public ZIP-level LTL coverage", "Asset-backed broker using Amazon logistics assets and vetted carrier network; not disclosed as dedicated LTL terminal network", "Online portal and instant quotes for eligible freight; LTL not yet a broad public product", "FTL portal, tracking, support, transport consultants for small/mid-sized inbound-to-Amazon customers", "AMZN26_PR; AMZN24_FAQ; AMZN_HOME"],
    ["2025-04-04", "Inbound LTL launch", "Customers shipping inbound to Amazon facilities", "Amazon fulfillment centers only", "Thousands of lanes coast-to-coast for inbound-to-Amazon use case", "70,000+ trailers referenced in launch; dedicated LTL operations support", "Self-service portal, quote up to 14 days out, compare LTL and FTL options", "Quote, book, track, billing/invoicing/payments through portal", "AMZN25_IN"],
    ["2025-09-10", "Inbound LTL booking workflow", "Logged-in Amazon Freight users", "Amazon FC destination selected by facility code", "Workflow confirms lane quoting inside account-gated portal", "Uses Amazon Freight LTL operations and facility-code routing", "Portal estimated quotes after origin, FC, pallet, PO and class inputs", "Pickup calendar, pallet details, freight class, shipment info and booking steps", "AMZN25_BOOK"],
    ["2026-06-10", "General-purpose LTL launch", "All business sizes, from smaller shippers to enterprise", "Amazon FCs plus third-party warehouses, distribution centers, retail partners and own facilities", "Public launch says destinations nationwide; coverage map still governs observed pickup/delivery availability", "80,000 trailers, 24,000 intermodal containers and terminals across major U.S. metros support broader freight platform", "Create account for quotes; enterprise sales contact", "One to six pallets / 150-15,000 lb, next-day live pickup, drop trailer, daily pickup, GPS, milestones, ePOD, EDI", "AMZN26_PR; AMZN26_LTL"],
    ["2026 current", "Public LTL page and coverage map", "Businesses of all sizes; SMB self-service and enterprise sales path", "Any type of destination where LTL pickup/delivery coverage is available", "Map-defined pickup, delivery and Amazon-destination coverage; more metros coming in 2026", "Amazon Freight is an asset-backed broker with Amazon-owned trailers and vetted carriers; LTL-specific terminals/assets not separately quantified", "Account-gated online quotes; web portal / EDI / API / TMS integrations", "Compare rates, book, manage freight, GPS tracking and milestone visibility", "AMZN26_PAGE; AMZN_HOME"],
]


customer_overlap = [
    ["Company size", "SMB and middle-market shippers", "FedEx Freight explicitly targets high-margin SMB growth with a dedicated LTL salesforce; Amazon says SMBs can self-serve through an online portal and serves businesses of all sizes.", 1.0, "AMZN24_FAQ; AMZN26_PR; FDXF_INV"],
    ["Shipment size", "1-6 pallets / 150-15,000 lb", "Amazon's published LTL profile is squarely in the LTL small shipment band; FedEx Freight is a national LTL carrier serving this core use case.", 1.0, "AMZN26_PR; FDXF_FORM10"],
    ["Digital buying journey", "Instant quotes, portal booking, tracking, billing", "Amazon requires account creation but positions online rate comparison/booking/tracking as central; FedEx Freight is improving quote-to-bill transparency and digital freight tools.", 1.0, "AMZN25_IN; AMZN25_BOOK; AMZN_HOME; FDXF_INV"],
    ["Destination type", "Warehouses, DCs, retail partners, Amazon FCs", "Amazon expanded from Amazon FC inbound to general warehouses/DCs/retail destinations; this overlaps with SMB distribution use cases FedEx Freight wants.", 0.8, "AMZN26_PR; AMZN26_LTL; FDXF_INV"],
    ["Service promise", "Reliable national LTL with pickup options", "Amazon offers next-day live pickup after 5 p.m. cutoff, drop trailer and milestone visibility; FedEx Freight offers established Priority/Economy service design.", 0.6, "AMZN26_PR; FDXF_INV"],
    ["Price-sensitive freight", "Transactional SMB/digital freight", "Both propositions emphasize transparent pricing; Amazon's price advantage cannot be proven without account-gated quotes.", 0.7, "AMZN_HOME; FDXF_INV"],
    ["Geographic overlap", "Dense major metros and linehaul corridors", "Amazon appears present in many major freight metros, including many FedEx Freight top-terminal markets, but lacks a public ZIP-level LTL matrix.", 0.7, "AMZN26_PAGE; FDXF_FORM10"],
    ["Freight type exclusions", "General dry freight, not specialized national LTL breadth", "Amazon Freight is asset-backed brokerage using Amazon assets and partner carriers; FedEx Freight is a dedicated LTL operating company with >365 locations.", 0.4, "AMZN_HOME; FDXF_FORM10"],
]


coverage = [
    ["Dallas/Fort Worth", "TX", "YES", "Irving top terminal", 277, "YES", "YES", "YES", "YES", "YES", "YES", "Dense Amazon map circle; major FDXF terminal"],
    ["Houston", "TX", "YES", "Houston top terminal", 241, "YES", "YES", "YES", "YES", "YES", "YES", "Dense Amazon map circle; major FDXF terminal"],
    ["San Antonio/Austin", "TX", "YES", "Major Texas market", None, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows central Texas coverage"],
    ["Los Angeles/Inland Empire", "CA", "YES", "Mira Loma and San Bernardino top terminals", 527, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Southern California coverage"],
    ["San Francisco/Oakland", "CA", "YES", "Bay Area market", None, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map-derived Bay Area/Northern California coverage"],
    ["Sacramento", "CA", "YES", "Northern California market", None, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map-derived Sacramento coverage"],
    ["Seattle", "WA", "YES", "Pacific Northwest market", None, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Seattle coverage"],
    ["Portland", "OR", "YES", "Pacific Northwest market", None, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Portland coverage"],
    ["Salt Lake City", "UT", "YES", "Mountain West market", None, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Salt Lake City coverage"],
    ["Las Vegas", "NV", "YES", "Southwest market", None, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Las Vegas coverage"],
    ["Chicago", "IL", "YES", "Chicago Heights and Forest View top terminals", 497, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Chicago/Milwaukee cluster"],
    ["Milwaukee", "WI", "YES", "Upper Midwest market", None, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Milwaukee coverage"],
    ["Minneapolis/St. Paul", "MN", "YES", "Lakeville top terminal", 234, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Minneapolis/St. Paul coverage"],
    ["Detroit", "MI", "YES", "Great Lakes market", None, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Detroit coverage"],
    ["Cleveland/North Jackson", "OH", "YES", "North Jackson top terminal", 235, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Cleveland/Pittsburgh corridor"],
    ["Columbus", "OH", "YES", "West Jefferson top terminal", 256, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows central Ohio coverage"],
    ["Cincinnati/Dayton", "OH", "YES", "Huber Heights top terminal", 227, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Cincinnati/Dayton corridor"],
    ["Indianapolis", "IN", "YES", "Indianapolis top terminal", 252, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Indianapolis coverage"],
    ["Louisville", "KY", "YES", "Louisville top terminal", 223, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Louisville coverage"],
    ["Nashville", "TN", "YES", "Southeast market", None, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Nashville coverage"],
    ["Memphis", "TN", "YES", "Southeast freight hub", None, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Memphis coverage"],
    ["Kansas City", "KS/MO", "YES", "Edwardsville top terminal", 339, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Kansas City coverage"],
    ["St. Louis", "MO", "YES", "St. Charles top terminal", 289, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows St. Louis coverage"],
    ["Atlanta", "GA", "YES", "Conley top terminal", 230, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Atlanta coverage"],
    ["Charlotte", "NC", "YES", "Charlotte top terminal", 227, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Charlotte coverage"],
    ["Raleigh", "NC", "YES", "Carolinas market", None, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Raleigh coverage"],
    ["Jacksonville", "FL", "YES", "Florida market", None, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Florida coverage"],
    ["Orlando/Tampa", "FL", "YES", "Florida market", None, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Orlando/Tampa coverage"],
    ["Miami", "FL", "YES", "South Florida market", None, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Miami coverage"],
    ["New York/Northern New Jersey", "NY/NJ", "YES", "Northeast market", None, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows dense Northeast coverage"],
    ["Philadelphia", "PA", "YES", "Middletown top terminal", 241, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Philadelphia coverage"],
    ["Baltimore/DC/Hagerstown", "MD/DC", "YES", "Hagerstown top terminal", 284, "YES", "YES", "YES", "YES", "YES", "YES", "Amazon map shows Mid-Atlantic coverage"],
    ["Boston", "MA", "YES", "New England market", None, "NO", "PARTIAL", "NO", "YES", "YES", "YES", "Amazon map has Amazon-delivery marker nearby, but not clear LTL pickup/delivery circle"],
    ["Denver", "CO", "YES", "Henderson top terminal", 241, "NO", "NO", "NO", "YES", "YES", "YES", "Major FDXF terminal, not evident in Amazon LTL map"],
    ["Phoenix", "AZ", "YES", "Southwest market", None, "NO", "NO", "NO", "YES", "YES", "YES", "Major market, not evident in Amazon LTL map"],
    ["Des Moines", "IA", "YES", "Des Moines top terminal", 222, "NO", "NO", "NO", "YES", "YES", "YES", "Top FDXF terminal, not evident in Amazon LTL map"],
]


lanes = [
    ("L001", "Houston", "Dallas/Fort Worth", "Regional TX"),
    ("L002", "Dallas/Fort Worth", "Chicago", "Long-haul core"),
    ("L003", "Atlanta", "Charlotte", "Southeast"),
    ("L004", "New York/Northern New Jersey", "Boston", "Northeast partial"),
    ("L005", "Chicago", "Nashville", "Midwest-South"),
    ("L006", "Los Angeles/Inland Empire", "Phoenix", "Southwest partial"),
    ("L007", "Los Angeles/Inland Empire", "Dallas/Fort Worth", "Long-haul core"),
    ("L008", "Houston", "Atlanta", "Gulf-Southeast"),
    ("L009", "Chicago", "Detroit", "Great Lakes"),
    ("L010", "Atlanta", "Miami", "Southeast-Florida"),
    ("L011", "Dallas/Fort Worth", "Houston", "Regional TX"),
    ("L012", "Dallas/Fort Worth", "San Antonio/Austin", "Regional TX"),
    ("L013", "Houston", "San Antonio/Austin", "Regional TX"),
    ("L014", "Los Angeles/Inland Empire", "Las Vegas", "Southwest"),
    ("L015", "Los Angeles/Inland Empire", "Salt Lake City", "West"),
    ("L016", "Seattle", "Portland", "Pacific Northwest"),
    ("L017", "Portland", "Sacramento", "West Coast"),
    ("L018", "San Francisco/Oakland", "Los Angeles/Inland Empire", "California"),
    ("L019", "Sacramento", "Las Vegas", "West"),
    ("L020", "Chicago", "Indianapolis", "Midwest"),
    ("L021", "Chicago", "Cleveland/North Jackson", "Great Lakes"),
    ("L022", "Cleveland/North Jackson", "Pittsburgh", "Great Lakes"),
    ("L023", "Detroit", "Cleveland/North Jackson", "Great Lakes"),
    ("L024", "Columbus", "Cincinnati/Dayton", "Ohio"),
    ("L025", "Cincinnati/Dayton", "Louisville", "Ohio Valley"),
    ("L026", "Louisville", "Nashville", "South"),
    ("L027", "Nashville", "Memphis", "Tennessee"),
    ("L028", "Memphis", "Atlanta", "Southeast"),
    ("L029", "Kansas City", "St. Louis", "Missouri"),
    ("L030", "Kansas City", "Dallas/Fort Worth", "Central-South"),
    ("L031", "St. Louis", "Chicago", "Midwest"),
    ("L032", "Minneapolis/St. Paul", "Chicago", "Upper Midwest"),
    ("L033", "Minneapolis/St. Paul", "Milwaukee", "Upper Midwest"),
    ("L034", "Milwaukee", "Detroit", "Great Lakes"),
    ("L035", "Atlanta", "Jacksonville", "Southeast-Florida"),
    ("L036", "Jacksonville", "Orlando/Tampa", "Florida"),
    ("L037", "Orlando/Tampa", "Miami", "Florida"),
    ("L038", "Charlotte", "Raleigh", "Carolinas"),
    ("L039", "Raleigh", "Atlanta", "Southeast"),
    ("L040", "New York/Northern New Jersey", "Philadelphia", "Northeast"),
    ("L041", "Philadelphia", "Baltimore/DC/Hagerstown", "Mid-Atlantic"),
    ("L042", "Baltimore/DC/Hagerstown", "Charlotte", "Mid-Atlantic-South"),
    ("L043", "Cleveland/North Jackson", "Philadelphia", "Great Lakes-East"),
    ("L044", "Denver", "Kansas City", "Mountain-Central partial"),
    ("L045", "Denver", "Salt Lake City", "Mountain partial"),
    ("L046", "Phoenix", "Los Angeles/Inland Empire", "Southwest partial"),
    ("L047", "Phoenix", "Dallas/Fort Worth", "Southwest partial"),
    ("L048", "Des Moines", "Chicago", "Midwest partial"),
    ("L049", "Boston", "Philadelphia", "Northeast partial"),
    ("L050", "St. Louis", "Nashville", "Midwest-South"),
]


profiles = [
    ("A", "SMALL SMB PALLET: 2 pallets, 1,000 lb total, standard dimensions, standard freight class, business-to-business"),
    ("B", "MID-SIZE LTL: 6 pallets, 4,000 lb total, standard business-to-business LTL"),
    ("C", "HIGHER-VALUE PALLETIZED GOODS: 2 pallets, 1,500 lb total, higher freight class"),
]


service_comparison = [
    ["General-purpose LTL availability", "Available nationally through mature LTL network", "Launched broadly on Jun. 10, 2026 with map-defined pickup/delivery zones", "ADVANTAGE - FDXF", "Amazon is expanding quickly, but FDXF has an established national LTL operating base.", "AMZN26_PR; FDXF_FORM10"],
    ["Digital quote and booking", "Quote tools and enhanced quote-to-bill transparency", "Account-gated SMB instant quotes and portal booking", "PARITY", "Both target digital self-service; direct quotes require shipper details/account context.", "AMZN_HOME; AMZN25_BOOK; FDXF_INV"],
    ["Shipment tracking", "FedEx tracking and LTL tools", "Real-time GPS, milestone visibility, ePOD and sensor-based visibility claims", "ADVANTAGE - Amazon", "Amazon's public copy is unusually explicit about trailer/pallet visibility.", "AMZN26_PR; AMZN26_LTL"],
    ["Service product breadth", "Priority/Economy dual offering; specialized Freight Direct and Custom Critical adjacencies", "General LTL plus FTL/intermodal freight platform", "ADVANTAGE - FDXF", "FDXF publishes a purpose-built LTL service set; Amazon's LTL product is newer and less segmented publicly.", "FDXF_INV; FDXF_PAGE; AMZN_HOME"],
    ["Coverage transparency", "Large disclosed terminal network and service to nearly every U.S. ZIP", "Public coverage map with pickup/delivery circles and more metros coming in 2026", "ADVANTAGE - FDXF", "Amazon provides visual coverage but not a structured ZIP-to-ZIP matrix.", "FDXF_FORM10; AMZN26_PAGE"],
    ["Drop-trailer convenience", "Available to eligible LTL/customers but not the core public proof point", "Unified drop-trailer pool and daily pickup programs highlighted", "ADVANTAGE - Amazon", "Amazon is leveraging existing Amazon trailer density as a convenience feature.", "AMZN26_PR; AMZN26_LTL"],
    ["Claims/on-time proof", "Mature LTL operating metrics and established claims process", "Claims and LTL on-time performance not yet publicly proven", "UNKNOWN", "Lack of public Amazon LTL performance data contradicts a full-threat thesis.", "FDXF_FORM10; AMZN26_PR"],
    ["SMB sales motion", "500+ dedicated LTL sales team plus SMB growth priority", "SMB online portal; enterprise sales contact", "PARITY", "Both are trying to reduce friction for smaller shippers.", "FDXF_INV; AMZN24_FAQ; AMZN_HOME"],
]


local_files = {
    "AMZN26_PR": "amazon_2026_ltl_press_release.html",
    "AMZN26_LTL": "amazon_2026_ltl_launch.html",
    "AMZN26_PAGE": "amazon_2026_ltl_service_page.html; amazon_ltl_coverage_map.pdf",
    "AMZN25_IN": "amazon_2025_ltl_inbound_launch.html",
    "AMZN25_BOOK": "amazon_2025_ltl_booking.html",
    "AMZN24_FAQ": "amazon_2024_faq.html",
    "AMZN_HOME": "amazon_freight_home.html",
    "FDXF_FORM10": "fdxf_form10_ex99_1.html",
    "FDXF_INV": "../phase2a_FedEx_Freight_Investor_Day_Presentation.pdf",
    "FDXF_PAGE": "not archived; public page used for context",
    "XPO_PAGE": "xpo_geographic_coverage.html",
    "ODFL_PAGE": "odfl_about.html; odfl_service_center_locator.html",
    "SAIA_10K": "saia_2025_10k.html",
}


def covered_market_set():
    return {row[0] for row in coverage if row[7] == "YES"}


def lane_distance_bucket(origin, dest):
    short_pairs = {
        frozenset(["Houston", "Dallas/Fort Worth"]),
        frozenset(["Dallas/Fort Worth", "San Antonio/Austin"]),
        frozenset(["Houston", "San Antonio/Austin"]),
        frozenset(["Seattle", "Portland"]),
        frozenset(["Chicago", "Indianapolis"]),
        frozenset(["Columbus", "Cincinnati/Dayton"]),
        frozenset(["Cincinnati/Dayton", "Louisville"]),
        frozenset(["Jacksonville", "Orlando/Tampa"]),
        frozenset(["Charlotte", "Raleigh"]),
        frozenset(["Philadelphia", "Baltimore/DC/Hagerstown"]),
    }
    long_pairs = {
        frozenset(["Dallas/Fort Worth", "Chicago"]),
        frozenset(["Los Angeles/Inland Empire", "Dallas/Fort Worth"]),
        frozenset(["Houston", "Atlanta"]),
        frozenset(["Atlanta", "Miami"]),
        frozenset(["Los Angeles/Inland Empire", "Salt Lake City"]),
        frozenset(["Denver", "Salt Lake City"]),
        frozenset(["Phoenix", "Dallas/Fort Worth"]),
    }
    key = frozenset([origin, dest])
    if key in short_pairs:
        return "<300 mi"
    if key in long_pairs:
        return ">1,000 mi"
    return "300-1,000 mi"


def table(ws, name, last_row, last_col):
    ref = f"A1:{get_column_letter(last_col)}{last_row}"
    tab = Table(displayName=name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)


def style_ws(ws, widths=None, freeze="A2"):
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if widths:
        for col, width in widths.items():
            ws.column_dimensions[col].width = width


def append_sheet(wb, title, headers, rows, widths=None, table_name=None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_ws(ws, widths or {})
    if table_name and rows:
        table(ws, table_name, len(rows) + 1, len(headers))
    return ws


def build():
    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    summary = wb.create_sheet("Executive_Summary")
    summary.append(["Question / Metric", "Answer / Formula", "Evidence Basis", "Source IDs"])
    summary_rows = [
        ["1. How rapidly has Amazon expanded LTL?", "Rapid but recent: 2025 inbound-to-Amazon FC LTL, Sep. 2025 account-gated booking workflow, Jun. 2026 general-purpose LTL for all businesses and destination types.", "Timeline shows a material 14-month progression, but broad public LTL is only a 2026 launch.", "AMZN25_IN; AMZN25_BOOK; AMZN26_PR; AMZN26_LTL"],
        ["2. How much geographic overlap exists with FDXF?", "=COUNTIF(Coverage!H2:H37,\"YES\")/COUNTA(Coverage!A2:A37)", "Unweighted major-market proxy, not ZIP-level or volume-weighted national coverage.", "AMZN26_PAGE; FDXF_FORM10"],
        ["FDXF top-terminal door-weighted Amazon overlap", "=SUMPRODUCT((Coverage!H2:H37=\"YES\")*(Coverage!E2:E37))/SUM(Coverage!E2:E37)", "Uses disclosed FedEx Freight top-terminal door counts where applicable; empty door cells are excluded.", "FDXF_FORM10; AMZN26_PAGE"],
        ["3. Does Amazon target the same SMB/digital customer FDXF wants?", "=AVERAGE(Customer_Overlap!D2:D9)", "High overlap in SMB, digital quoting, palletized LTL and warehouse/DC/retail destinations; lower where Amazon lacks mature published LTL breadth.", "AMZN_HOME; AMZN26_PR; FDXF_INV"],
        ["4. Are direct quotes testable publicly?", "=SUMPRODUCT(--(COUNTIFS(Quotes_Raw!A2:A751,Test_Lanes!A2:A51,Quotes_Raw!I2:I751,\"Amazon Freight\",Quotes_Raw!J2:J751,\"YES\")>0))/COUNTA(Test_Lanes!A2:A51)", "No comparable public quotes were captured because Amazon and peer quote tools require account/shipment credentials.", "AMZN25_BOOK; AMZN_HOME"],
        ["Amazon map-covered test lane share", "=COUNTIF(Test_Lanes!I2:I51,\"YES\")/COUNTA(Test_Lanes!A2:A51)", "Lane coverage is inferred only from Amazon's public LTL coverage map.", "AMZN26_PAGE"],
        ["5. Is Amazon material or emerging?", "Emerging competitor for SMB/digital freight in covered metros, not yet proven as a full national LTL substitute for FedEx Freight.", "Material overlap in dense metros and customer motion; weaker evidence on prices, service consistency, claims, ZIP-level reach and LTL operating depth.", "AMZN26_PR; FDXF_FORM10; FDXF_INV"],
        ["6. Evidence contradicting bearish thesis", "Amazon is asset-backed brokerage, not disclosed as a nationwide dedicated LTL terminal network; map coverage is partial/visual; public prices and transit times are not observable; FDXF has >365 locations and service to nearly every U.S. ZIP.", "These points prevent a forced major-threat conclusion.", "AMZN_HOME; AMZN26_PAGE; FDXF_FORM10"],
    ]
    for row in summary_rows:
        summary.append(row)
    style_ws(summary, {"A": 36, "B": 62, "C": 52, "D": 28})
    for cell in ["B3", "B4", "B5", "B6", "B7"]:
        summary[cell].number_format = "0.0%"

    ws = append_sheet(
        wb,
        "Amazon_Timeline",
        ["date", "product_scope", "eligible_shipper", "eligible_destination", "geographic_coverage", "network_claim", "pricing_method", "service_features", "source_url", "source_ids"],
        [row[:-1] + [src_urls(row[-1]), row[-1]] for row in timeline],
        {"A": 16, "B": 34, "C": 44, "D": 44, "E": 54, "F": 62, "G": 44, "H": 58, "I": 74, "J": 26},
        "AmazonTimeline",
    )

    co_rows = []
    for dim, theme, evidence, score, src in customer_overlap:
        co_rows.append([dim, theme, evidence, score, score_label(score), src])
    ws = append_sheet(
        wb,
        "Customer_Overlap",
        ["Dimension", "Overlap Theme", "Evidence", "Numeric Score", "Score Label", "Source IDs"],
        co_rows,
        {"A": 24, "B": 34, "C": 78, "D": 14, "E": 14, "F": 28},
        "CustomerOverlap",
    )
    for row in range(2, len(co_rows) + 2):
        ws[f"D{row}"].number_format = "0.0"
    ws.conditional_formatting.add(f"D2:D{len(co_rows)+1}", CellIsRule(operator="greaterThanOrEqual", formula=["0.8"], fill=PatternFill("solid", fgColor="C6EFCE")))
    ws.conditional_formatting.add(f"D2:D{len(co_rows)+1}", CellIsRule(operator="between", formula=["0.5", "0.799"], fill=PatternFill("solid", fgColor="FFEB9C")))

    cov_rows = []
    for row in coverage:
        cov_rows.append(row + [f'=IF(H{len(cov_rows)+2}="YES",1,0)', "AMZN26_PAGE; FDXF_FORM10; ODFL_PAGE; XPO_PAGE; SAIA_10K"])
    ws = append_sheet(
        wb,
        "Coverage",
        ["Market", "State", "FDXF Presence", "FedEx Basis", "FDXF Top-Terminal Doors", "Amazon Pickup", "Amazon Delivery", "Amazon LTL Available", "ODFL Presence", "XPO Presence", "Saia Presence", "Coverage Note", "Amazon Available Numeric", "Source IDs"],
        cov_rows,
        {"A": 28, "B": 12, "C": 14, "D": 44, "E": 16, "F": 14, "G": 14, "H": 16, "I": 14, "J": 14, "K": 14, "L": 58, "M": 16, "N": 34},
        "CoverageProxy",
    )
    for row in range(2, len(cov_rows) + 2):
        ws[f"E{row}"].number_format = "0"

    covered = covered_market_set()
    lane_rows = []
    for lane_id, origin, dest, region in lanes:
        cov = "YES" if origin in covered and dest in covered else "NO"
        lane_rows.append([lane_id, origin, dest, region, lane_distance_bucket(origin, dest), "A/B/C", "YES" if origin in covered else "NO", "YES" if dest in covered else "NO", cov, REGION.get(origin, "Other"), REGION.get(dest, "Other"), "Coverage from Amazon public map; no ZIP-level validation", "AMZN26_PAGE"])
    ws = append_sheet(
        wb,
        "Test_Lanes",
        ["Lane ID", "Origin Market", "Destination Market", "Region / Corridor", "Distance Bucket", "Shipment Profiles Tested", "Origin Amazon Covered", "Destination Amazon Covered", "Amazon Map-Covered Lane", "Origin Region", "Destination Region", "Lane Note", "Source IDs"],
        lane_rows,
        {"A": 10, "B": 28, "C": 28, "D": 24, "E": 14, "F": 20, "G": 18, "H": 22, "I": 20, "J": 18, "K": 20, "L": 52, "M": 18},
        "TestLanes",
    )

    carriers = ["Amazon Freight", "FedEx Freight", "XPO", "Old Dominion", "Saia"]
    quote_rows = []
    for lane_id, origin, dest, region in lanes:
        map_covered = "YES" if origin in covered and dest in covered else "NO"
        dist = lane_distance_bucket(origin, dest)
        for profile_code, profile_desc in profiles:
            for carrier in carriers:
                if carrier == "Amazon Freight":
                    status = "MAP_COVERED_NOT_PRICE_TESTED" if map_covered == "YES" else "MAP_NOT_COVERED"
                    reason = "Amazon public quote workflow requires account creation/login and shipment details; no credentialed quote collected."
                    src = "AMZN25_BOOK; AMZN_HOME; AMZN26_PAGE"
                else:
                    status = "PUBLIC_QUOTE_NOT_COLLECTED"
                    reason = "Comparable peer quote collection requires account/rate context or detailed shipper information; automated quote scraping was not attempted."
                    src = "FDXF_PAGE; XPO_PAGE; ODFL_PAGE; SAIA_10K"
                quote_rows.append([lane_id, origin, dest, REGION.get(origin, "Other"), REGION.get(dest, "Other"), dist, profile_code, profile_desc, carrier, "NO", None, None, None, None, status, reason, src])
    ws = append_sheet(
        wb,
        "Quotes_Raw",
        ["Lane ID", "Origin Market", "Destination Market", "Origin Region", "Destination Region", "Distance Bucket", "Profile Code", "Profile Description", "Carrier", "Quote Available Publicly", "Quoted Price USD", "Transit Days", "Pickup Lead Time", "Accessorials", "Status", "Status Reason", "Source IDs"],
        quote_rows,
        {"A": 10, "B": 26, "C": 26, "D": 18, "E": 20, "F": 16, "G": 12, "H": 64, "I": 18, "J": 20, "K": 16, "L": 14, "M": 18, "N": 24, "O": 30, "P": 72, "Q": 28},
        "QuotesRaw",
    )

    qa_rows = [
        ["Total test lanes", "=COUNTA(Test_Lanes!A2:A51)", "50 representative SMB/digital LTL corridors", "Test_Lanes"],
        ["Amazon map-covered lanes", "=COUNTIF(Test_Lanes!I2:I51,\"YES\")", "Both endpoints fall inside Amazon public LTL coverage proxy", "Test_Lanes"],
        ["Amazon map coverage share", "=B3/B2", "Lane-level map overlap proxy", "Test_Lanes"],
        ["Total raw quote attempts", "=COUNTA(Quotes_Raw!A2:A751)", "50 lanes x 3 shipment profiles x 5 carriers", "Quotes_Raw"],
        ["Public comparable quotes captured", "=COUNTIF(Quotes_Raw!J2:J751,\"YES\")", "Account-gated tools prevented directly comparable public quotes", "Quotes_Raw"],
        ["Public quote capture rate", "=B6/B5", "Observed public quote availability, not market competitiveness", "Quotes_Raw"],
        ["Amazon quote coverage rate - lanes quoted / lanes tested", "=SUMPRODUCT(--(COUNTIFS(Quotes_Raw!A2:A751,Test_Lanes!A2:A51,Quotes_Raw!I2:I751,\"Amazon Freight\",Quotes_Raw!J2:J751,\"YES\")>0))/COUNTA(Test_Lanes!A2:A51)", "Direct lane-level Amazon quote rate; zero because the public interface is account-gated.", "Quotes_Raw"],
        ["Profile A Amazon public quote capture", "=COUNTIFS(Quotes_Raw!I2:I751,\"Amazon Freight\",Quotes_Raw!G2:G751,\"A\",Quotes_Raw!J2:J751,\"YES\")/COUNTIFS(Quotes_Raw!I2:I751,\"Amazon Freight\",Quotes_Raw!G2:G751,\"A\")", "Small SMB pallet profile from the brief.", "Quotes_Raw"],
        ["Profile B Amazon public quote capture", "=COUNTIFS(Quotes_Raw!I2:I751,\"Amazon Freight\",Quotes_Raw!G2:G751,\"B\",Quotes_Raw!J2:J751,\"YES\")/COUNTIFS(Quotes_Raw!I2:I751,\"Amazon Freight\",Quotes_Raw!G2:G751,\"B\")", "Mid-size LTL profile from the brief.", "Quotes_Raw"],
        ["Profile C Amazon public quote capture", "=COUNTIFS(Quotes_Raw!I2:I751,\"Amazon Freight\",Quotes_Raw!G2:G751,\"C\",Quotes_Raw!J2:J751,\"YES\")/COUNTIFS(Quotes_Raw!I2:I751,\"Amazon Freight\",Quotes_Raw!G2:G751,\"C\")", "Higher-value palletized goods profile from the brief.", "Quotes_Raw"],
        ["Map coverage by distance: <300 mi", "=COUNTIFS(Test_Lanes!E2:E51,\"<300 mi\",Test_Lanes!I2:I51,\"YES\")/COUNTIF(Test_Lanes!E2:E51,\"<300 mi\")", "Coverage proxy by distance bucket.", "Test_Lanes"],
        ["Map coverage by distance: 300-1,000 mi", "=COUNTIFS(Test_Lanes!E2:E51,\"300-1,000 mi\",Test_Lanes!I2:I51,\"YES\")/COUNTIF(Test_Lanes!E2:E51,\"300-1,000 mi\")", "Coverage proxy by distance bucket.", "Test_Lanes"],
        ["Map coverage by distance: >1,000 mi", "=COUNTIFS(Test_Lanes!E2:E51,\">1,000 mi\",Test_Lanes!I2:I51,\"YES\")/COUNTIF(Test_Lanes!E2:E51,\">1,000 mi\")", "Coverage proxy by distance bucket.", "Test_Lanes"],
        ["Origin region map coverage: Texas", "=COUNTIFS(Test_Lanes!J2:J51,\"Texas\",Test_Lanes!I2:I51,\"YES\")/COUNTIF(Test_Lanes!J2:J51,\"Texas\")", "Origin-region directional coverage proxy.", "Test_Lanes"],
        ["Origin region map coverage: West", "=COUNTIFS(Test_Lanes!J2:J51,\"West\",Test_Lanes!I2:I51,\"YES\")/COUNTIF(Test_Lanes!J2:J51,\"West\")", "Origin-region directional coverage proxy.", "Test_Lanes"],
        ["Origin region map coverage: Midwest", "=COUNTIFS(Test_Lanes!J2:J51,\"Midwest\",Test_Lanes!I2:I51,\"YES\")/COUNTIF(Test_Lanes!J2:J51,\"Midwest\")", "Origin-region directional coverage proxy.", "Test_Lanes"],
        ["Origin region map coverage: Southeast", "=COUNTIFS(Test_Lanes!J2:J51,\"Southeast\",Test_Lanes!I2:I51,\"YES\")/COUNTIF(Test_Lanes!J2:J51,\"Southeast\")", "Origin-region directional coverage proxy.", "Test_Lanes"],
        ["Destination region map coverage: Northeast", "=COUNTIFS(Test_Lanes!K2:K51,\"Northeast\",Test_Lanes!I2:I51,\"YES\")/COUNTIF(Test_Lanes!K2:K51,\"Northeast\")", "Destination-region directional coverage proxy.", "Test_Lanes"],
        ["Destination region map coverage: Florida", "=COUNTIFS(Test_Lanes!K2:K51,\"Florida\",Test_Lanes!I2:I51,\"YES\")/COUNTIF(Test_Lanes!K2:K51,\"Florida\")", "Destination-region directional coverage proxy.", "Test_Lanes"],
        ["Amazon direct price advantage", "N/A", "Cannot compute without public or credentialed Amazon quotes", "Quotes_Raw"],
        ["Transit-time comparison", "N/A", "Amazon route-specific transit times were not public in collected sources", "Quotes_Raw"],
        ["Convenience comparison", "Amazon positive on portal/GPS/drop-trailer; FDXF positive on mature LTL network/service breadth", "Qualitative service evidence only", "Service_Comparison"],
    ]
    ws = append_sheet(
        wb,
        "Quote_Analysis",
        ["Metric", "Value / Formula", "Interpretation", "Source Sheet"],
        qa_rows,
        {"A": 34, "B": 28, "C": 70, "D": 24},
        "QuoteAnalysis",
    )
    for cell in ["B4", "B7", "B8", "B9", "B10", "B11", "B12", "B13", "B14", "B15", "B16", "B17", "B18", "B19"]:
        ws[cell].number_format = "0.0%"

    append_sheet(
        wb,
        "Service_Comparison",
        ["Dimension", "FedEx Freight Evidence", "Amazon Freight Evidence", "Relative Read", "Interpretation", "Source IDs"],
        service_comparison,
        {"A": 28, "B": 48, "C": 50, "D": 18, "E": 62, "F": 26},
        "ServiceComparison",
    )

    source_rows = []
    for sid, title in SOURCE.items():
        source_rows.append([sid, title, URL[sid], local_files.get(sid, ""), "Primary/company or SEC source" if sid != "XPO_PAGE" and sid != "ODFL_PAGE" else "Official company source"])
    append_sheet(
        wb,
        "Sources",
        ["Source ID", "Description", "URL", "Local Archive File(s)", "Source Type"],
        source_rows,
        {"A": 16, "B": 68, "C": 92, "D": 48, "E": 28},
        "SourcesList",
    )

    limitations = [
        ["No direct public rate quotes", "Amazon Freight LTL pricing requires account creation/login and shipment details. Comparable public quote harvesting was not performed without credentials.", "High", "Price/transit competitiveness cannot be concluded from this workbook alone."],
        ["Amazon map is not ZIP-level data", "Amazon publishes visual LTL pickup/delivery/destination coverage. This workbook converts the map into a major-market proxy by manual map read.", "High", "Geographic overlap is directional, not a precise served-ZIP or volume-weighted estimate."],
        ["Amazon is asset-backed brokerage, not disclosed dedicated LTL network", "Amazon describes Amazon Freight as asset-backed brokerage supported by Amazon-owned trailers and vetted carriers.", "High", "Do not model Amazon as a full nationwide asset-based LTL carrier equivalent to FDXF."],
        ["FedEx Freight volume by market not public", "FDXF top-terminal doors are used as a proxy for dense-market exposure where disclosed.", "Medium", "Door-weighted overlap is a proxy, not revenue or shipment overlap."],
        ["Peer presence simplified", "ODFL, XPO and Saia are shown as present in major markets based on broad national networks and official locator/network disclosures.", "Medium", "Workbook is focused on Amazon-vs-FDXF threat, not a fully specified five-carrier routing model."],
        ["No operational performance history for Amazon LTL", "Amazon's general-purpose LTL launch is very recent, so lane-level on-time, claims, handling and damage metrics were not publicly available.", "High", "Service quality risk remains a key unknown and bearish-thesis counterweight."],
    ]
    append_sheet(
        wb,
        "Limitations",
        ["Limitation", "Explanation", "Severity", "Analytical Impact"],
        limitations,
        {"A": 34, "B": 72, "C": 14, "D": 58},
        "LimitationsList",
    )

    for ws in wb.worksheets:
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name="Aptos", size=10, bold=cell.row == 1, color="FFFFFF" if cell.row == 1 else "000000")
        if ws.title == "Executive_Summary":
            for cell in ws["A"][1:]:
                cell.font = Font(name="Aptos", size=10, bold=True)

    wb.save(WORKBOOK)

    # Basic read-back validation.
    loaded = load_workbook(WORKBOOK, data_only=False)
    expected = [
        "Executive_Summary",
        "Amazon_Timeline",
        "Customer_Overlap",
        "Coverage",
        "Test_Lanes",
        "Quotes_Raw",
        "Quote_Analysis",
        "Service_Comparison",
        "Sources",
        "Limitations",
    ]
    assert loaded.sheetnames == expected, loaded.sheetnames
    assert loaded["Test_Lanes"].max_row == 51
    assert loaded["Quotes_Raw"].max_row == 751
    assert loaded["Sources"].max_row == len(SOURCE) + 1
    print(WORKBOOK)
    print(f"sheets={len(loaded.sheetnames)} lanes=50 quote_rows=750 coverage_markets={len(coverage)}")


if __name__ == "__main__":
    build()
